#!/usr/bin/env python3
"""
Macrostrat ingestion: metadata/project validation + refs + columns + units
Assumption is that project exists in Macrostrat; project must be defined by valid existing project_id in metadata tab
There is currently no check that supplied project_id is correct.

Phase 3.0: schema introspection + declarative mapping extended through macrostrat.units,
while keeping entity-specific logic (identity, geometry semantics, joins, recomputed groupings) in plugins.

 - Mapping-driven SQL generation for:
   * macrostrat.refs
   * macrostrat.col_groups
   * macrostrat.cols
   * macrostrat.units
 - Schema guardrails for mapped entities and critical NOT NULL/no-default coverage
 - Plugin-driven logic for:
   * columns, refs, units
   * unit joins
   * sections / units_sections recompute

Requires:
- metadata tab of Excel template must have field "mapping_version" matching "mapping_version" in file macrostrat_mapping.json (contains declarative field mappings)

Notes:
- All target tables are schema-qualified under macrostrat.<table>.

Audit:
- Writes six files providing Excel-macrostrat key pairs and a metadata/action summary to ./audit:
  col_id_map_<project_id>_<timestamp>.tsv
  ref_id_map_<project_id>_<timestamp>.tsv
  unit_id_map_<project_id>_<timestamp>.tsv
  unit_strat_name_map_<project_id>_<timestamp>.tsv
  strat_name_lexicon_<project_id>_<timestamp>.tsv
  macrostrat_import_audit_<project_id>_<timestamp>.json

"""

import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import ROUND_FLOOR, ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any, Callable, Dict, List, Optional, Tuple

import psycopg2
import requests
from openpyxl import load_workbook

from macrostrat.core.database import Database, get_database
from macrostrat.core.exc import MacrostratError

# -----------------------------
# Engine utilities
# -----------------------------


def clean_cell_value(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def normalize_str(x: Any) -> Optional[str]:
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        return s if s else None
    s = str(x).strip()
    return s if s else None


def normalize_wkt(wkt: str) -> str:
    w = wkt.strip()
    w = re.sub(r"\s+", " ", w)
    return w


def read_sheet_as_dicts(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers: List[str] = []
    for h in rows[0]:
        h = clean_cell_value(h)
        if h is None:
            raise ColumnIngestionError(
                f"Sheet '{ws.title}' has empty header cells in first row."
            )
        headers.append(str(h).strip())

    out: List[Dict[str, Any]] = []
    for i, r in enumerate(rows[1:], start=2):
        if r is None:
            continue
        cleaned_row = [clean_cell_value(v) for v in r]
        if all(
            v is None or (isinstance(v, str) and v.strip() == "") for v in cleaned_row
        ):
            continue
        d = {
            headers[j]: (cleaned_row[j] if j < len(cleaned_row) else None)
            for j in range(len(headers))
        }
        d["_excel_row"] = i
        out.append(d)

    return headers, out


class ColumnIngestionError(MacrostratError):
    pass


def get_metadata_value(wb, sheet_name: str, key: str) -> Any:
    if sheet_name not in wb.sheetnames:
        raise ColumnIngestionError(f"Workbook does not contain a '{sheet_name}' tab.")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        k = clean_cell_value(row[0] if len(row) > 0 else None)
        v = clean_cell_value(row[1] if len(row) > 1 else None)
        if k == key:
            return v
    return None


def load_mapping(mapping_path: str) -> Dict[str, Any]:
    try:
        with open(mapping_path, "r", encoding="utf-8") as f:
            mapping = json.load(f)
    except FileNotFoundError:
        raise ColumnIngestionError(f"Mapping file not found: {mapping_path}")
    except json.JSONDecodeError as e:
        raise ColumnIngestionError(f"Mapping file is not valid JSON: {e}")

    for k in ("mapping_version", "sheets", "entities"):
        if k not in mapping:
            raise ColumnIngestionError(f"Mapping file missing required field '{k}'.")
    if not isinstance(mapping["entities"], dict) or not mapping["entities"]:
        raise ColumnIngestionError(
            "Mapping file 'entities' must be a non-empty object."
        )

    return mapping


# -----------------------------
# Audit artifacts
# -----------------------------


def utc_runstamp() -> Tuple[str, str]:
    """
    Returns (runstamp_compact, runstamp_iso_utc)
      - compact: YYYYMMDDTHHMMSSZ
      - iso:     YYYY-MM-DDTHH:MM:SSZ
    """
    dt = datetime.now(timezone.utc).replace(microsecond=0)
    iso = dt.isoformat().replace("00:00", "Z")
    compact = dt.strftime("%Y%m%dT%H%M%SZ")
    return compact, iso


def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def write_tsv_map(
    path: str, header_a: str, header_b: str, mapping: Dict[str, int]
) -> None:
    """
    Writes a 2-column TSV with a header row. Keys sorted for deterministic output.
    """
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(f"{header_a}\t{header_b}\n")
        for k in sorted(mapping.keys()):
            f.write(f"{k}\t{mapping[k]}\n")


def write_units_tsv_header_only(path: str, header_a: str, header_b: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(f"{header_a}\t{header_b}\n")


def write_audit_artifacts(
    audit_dir: str,
    runstamp_compact: str,
    runstamp_iso: str,
    script_name: str,
    project_id: int,
    mapping_version: str,
    excel_file: str,
    excel_sha256: str,
    ref_map: Dict[str, int],
    col_map: Dict[str, int],
    counts: Dict[str, Any],
    unit_map: Optional[Dict[str, int]] = None,
    units_excel_key_name: str = "excel_unit_id",
    units_db_key_name: str = "macrostrat_unit_id",
    strat_name_lexicon_filename: Optional[str] = None,
    unit_strat_name_map_filename: Optional[str] = None,
) -> None:
    ensure_dir(audit_dir)

    manifest_path = os.path.join(
        audit_dir, f"macrostrat_import_audit_{project_id}_{runstamp_compact}.json"
    )
    refs_path = os.path.join(
        audit_dir, f"ref_id_map_{project_id}_{runstamp_compact}.tsv"
    )
    cols_path = os.path.join(
        audit_dir, f"col_id_map_{project_id}_{runstamp_compact}.tsv"
    )
    units_path = os.path.join(
        audit_dir, f"unit_id_map_{project_id}_{runstamp_compact}.tsv"
    )

    # TSVs
    write_tsv_map(refs_path, "excel_ref_id", "macrostrat_ref_id", ref_map)
    write_tsv_map(cols_path, "excel_col_id", "macrostrat_col_id", col_map)
    if unit_map is None:
        write_units_tsv_header_only(units_path, units_excel_key_name, units_db_key_name)
    else:
        write_tsv_map(units_path, units_excel_key_name, units_db_key_name, unit_map)

    # Manifest (small)
    manifest = {
        "artifact_version": "1.0",
        "runstamp_utc": runstamp_iso,
        "project_id": project_id,
        "mapping_version": str(mapping_version).strip(),
        "script_name": script_name,
        "audit_dir": audit_dir,
        "excel_file": excel_file,
        "excel_sha256": excel_sha256,
        "counts": counts,
        "outputs": {
            "manifest_json": os.path.basename(manifest_path),
            "refs_tsv": os.path.basename(refs_path),
            "cols_tsv": os.path.basename(cols_path),
            "units_tsv": os.path.basename(units_path),
            "strat_name_lexicon_tsv": strat_name_lexicon_filename,
            "unit_strat_name_map_tsv": unit_strat_name_map_filename,
        },
    }
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, sort_keys=True)

    print(f"audit: wrote core artifacts to {audit_dir}")


# -----------------------------
# Engine: API + DB
# -----------------------------


def verify_project_id_via_api(project_id: int) -> str:
    api_url = f"https://macrostrat.org/api/defs/projects?project_id={project_id}"
    try:
        r = requests.get(api_url, timeout=10)
        r.raise_for_status()
        j = r.json()
    except Exception as e:
        raise ColumnIngestionError(f"API request failed: {e}")

    try:
        success_obj = j["success"]
        data_array = success_obj["data"]
        if not data_array:
            raise ColumnIngestionError("API returned no data for that project_id.")
        returned_project_id = data_array[0]["project_id"]
        returned_project_name = data_array[0]["project"]
    except (KeyError, IndexError, TypeError) as e:
        raise ColumnIngestionError(f"Unexpected API response structure: {e}")

    if returned_project_id != project_id:
        raise ColumnIngestionError(
            f"API returned project_id {returned_project_id}, expected {project_id}."
        )

    print(f"API OK: project_id {project_id} ('{returned_project_name}')")
    return returned_project_name


def connect_db(engine):
    url = engine.url
    try:
        conn = psycopg2.connect(
            dbname=url.database,
            user=url.username,
            password=url.password,
            host=url.host,
            port=url.port,
        )
        conn.autocommit = False
        return conn
    except psycopg2.Error as e:
        raise ColumnIngestionError(
            f"Could not connect to local Postgres '{dbname}': {e}"
        )


def verify_project_id_in_db(conn, project_id: int) -> None:
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM macrostrat.projects WHERE id = %s;", (project_id,))
        if cur.fetchone() is None:
            raise ColumnIngestionError(
                f"project_id {project_id} not found in macrostrat.projects.id (local DB)."
            )
    print(f"DB OK: project_id {project_id} exists in macrostrat.projects")


# -----------------------------
# Schema introspection
# -----------------------------


@dataclass(frozen=True)
class ColumnInfo:
    name: str
    data_type: str
    is_nullable: bool
    has_default: bool
    default_expr: Optional[str]


class SchemaInspector:
    def __init__(self, conn):
        self.conn = conn
        self._cols: Dict[Tuple[str, str], Dict[str, ColumnInfo]] = {}

    @staticmethod
    def split_table(fqtn: str) -> Tuple[str, str]:
        if "." not in fqtn:
            raise ColumnIngestionError(
                f"Expected schema-qualified table name, got '{fqtn}'."
            )
        schema, table = fqtn.split(".", 1)
        return schema, table

    def columns(self, fqtn: str) -> Dict[str, ColumnInfo]:
        schema, table = self.split_table(fqtn)
        key = (schema, table)
        if key in self._cols:
            return self._cols[key]

        with self.conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                  column_name,
                  data_type,
                  (is_nullable = 'YES') AS is_nullable,
                  (column_default IS NOT NULL) AS has_default,
                  column_default
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                ORDER BY ordinal_position
                """,
                (schema, table),
            )
            rows = cur.fetchall()

        if not rows:
            raise ColumnIngestionError(
                f"Could not introspect columns for table {schema}.{table} (not found?)."
            )

        cols: Dict[str, ColumnInfo] = {}
        for name, dtype, is_nullable, has_default, default_expr in rows:
            cols[name] = ColumnInfo(
                name=name,
                data_type=dtype,
                is_nullable=bool(is_nullable),
                has_default=bool(has_default),
                default_expr=default_expr,
            )

        self._cols[key] = cols
        return cols


def validate_mapping_against_schema(
    schema: SchemaInspector, mapping: Dict[str, Any]
) -> None:
    """
    Validation:
      - all mapped DB columns exist
      - mapped NOT NULL/no-default columns have a source/default in mapping
    """
    for ent_name, ent in mapping["entities"].items():
        fqtn = ent["table"]
        db_cols = schema.columns(fqtn)

        fields = ent.get("fields", {})
        if not isinstance(fields, dict) or not fields:
            raise ColumnIngestionError(
                f"Mapping entity '{ent_name}' must define non-empty 'fields'."
            )

        for db_col in fields.keys():
            if db_col not in db_cols:
                raise ColumnIngestionError(
                    f"Mapping entity '{ent_name}' refers to missing column '{db_col}' on {fqtn}."
                )

        for db_col, spec in fields.items():
            cinfo = db_cols[db_col]
            if not cinfo.is_nullable and not cinfo.has_default:
                has_source = any(
                    k in spec
                    for k in ("const", "sql", "expr", "from", "from_any", "handler")
                )
                has_default = "default" in spec
                if not (has_source or has_default):
                    raise ColumnIngestionError(
                        f"Entity '{ent_name}' column '{db_col}' is NOT NULL with no default in DB, "
                        f"but mapping provides no source/default."
                    )


def validate_cols_not_null_coverage(
    schema: SchemaInspector, cols_entity: Dict[str, Any]
) -> None:
    """
    Phase 2.5 safety guard:
      For macrostrat.cols specifically, ensure every NOT NULL column with no default
      is covered by the mapping (or is a DB default).
    """
    fqtn = cols_entity["table"]
    db_cols = schema.columns(fqtn)
    mapped = set(cols_entity.get("fields", {}).keys())

    missing_required: List[str] = []
    for name, info in db_cols.items():
        if info.is_nullable:
            continue
        if info.has_default:
            continue
        if name not in mapped:
            missing_required.append(name)

    if missing_required:
        raise ColumnIngestionError(
            "macrostrat.cols has NOT NULL columns with no defaults that are not mapped.\n"
            f"Missing in mapping: {missing_required}\n"
            "Update macrostrat_mapping.json (entity 'cols') to provide values/defaults."
        )


def validate_units_not_null_coverage(
    schema: SchemaInspector, units_entity: Dict[str, Any]
) -> None:
    """
    Phase 3 safety guard:
      For macrostrat.units specifically, ensure every NOT NULL column with no default
      is covered by the mapping (or is a DB default).
    """
    fqtn = units_entity["table"]
    db_cols = schema.columns(fqtn)
    mapped = set(units_entity.get("fields", {}).keys())

    missing_required: List[str] = []
    for name, info in db_cols.items():
        if info.is_nullable:
            continue
        if info.has_default:
            continue
        if name not in mapped:
            missing_required.append(name)

    if missing_required:
        raise ColumnIngestionError(
            "macrostrat.units has NOT NULL columns with no defaults that are not mapped.\n"
            f"Missing in mapping: {missing_required}\n"
            "Update macrostrat_mapping_v3.0.json (entity 'units') to provide values/defaults."
        )


# -----------------------------
# Mapping functions
# -----------------------------


def concat_ws_dot_space(title: Optional[str], publication: Optional[str]) -> str:
    parts = [p for p in [title, publication] if p is not None and str(p).strip() != ""]
    return ". ".join([str(p).strip() for p in parts])


EXPR_FUNCS = {"concat_ws_dot_space": concat_ws_dot_space}


def apply_transform(value: Any, transform: Optional[str]) -> Any:
    if transform is None:
        return value
    if transform == "lower":
        return value.lower() if isinstance(value, str) else value
    raise ColumnIngestionError(f"Unknown transform '{transform}'.")


def eval_field_spec(
    spec: Dict[str, Any],
    excel_row: Dict[str, Any],
    context: Dict[str, Any],
    handlers: Optional[
        Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]
    ] = None,
) -> Any:
    """
    Evaluate a mapping field spec to produce a value (python scalar OR SQL token/fragment dict).

    Supported in Phase 2.5:
      - const: "$project_id" or literal value
      - from: excel column name
      - from_any: [excel_col1, excel_col2, ...] first non-empty
      - expr: function name in EXPR_FUNCS with args (excel fields)
      - sql: SQL token, e.g. "now()"
      - handler: name resolved in handlers dict
      - transform: currently supports "lower"
      - default: used if value is None/empty string
    """
    value = None

    if "handler" in spec:
        if handlers is None:
            raise ColumnIngestionError(
                "Handler specified but no handlers registry provided."
            )
        hname = spec["handler"]
        hfn = handlers.get(hname)
        if hfn is None:
            raise ColumnIngestionError(f"Unknown handler '{hname}'.")
        value = hfn(excel_row, context)

    elif "const" in spec:
        c = spec["const"]
        if isinstance(c, str) and c.startswith("$"):
            key = c[1:]
            value = context.get(key)
        else:
            value = c

    elif "from" in spec:
        value = excel_row.get(spec["from"])

    elif "from_any" in spec:
        for k in spec["from_any"]:
            v = excel_row.get(k)
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                value = v
                break

    elif "expr" in spec:
        fn_name = spec["expr"]
        fn = EXPR_FUNCS.get(fn_name)
        if fn is None:
            raise ColumnIngestionError(f"Unknown expr function '{fn_name}'.")
        args = []
        for a in spec.get("args", []):
            args.append(excel_row.get(a))
        value = fn(*args)

    elif "sql" in spec:
        value = {"__sql__": spec["sql"]}

    else:
        raise ColumnIngestionError(f"Unsupported field spec: {spec}")

    value = apply_transform(value, spec.get("transform"))

    if value is None or (isinstance(value, str) and value.strip() == ""):
        if "default" in spec:
            value = spec["default"]

    if isinstance(value, str):
        value = value.strip()

    return value


def eval_entity_row_values(
    entity: Dict[str, Any],
    excel_row: Dict[str, Any],
    context: Dict[str, Any],
    handlers: Optional[
        Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]
    ] = None,
) -> Dict[str, Any]:
    fields = entity["fields"]
    row_values: Dict[str, Any] = {}
    for db_col, spec in fields.items():
        val = eval_field_spec(spec, excel_row, context, handlers=handlers)
        if spec.get("omit_if_none") and val is None:
            continue
        row_values[db_col] = val
    return row_values


def build_insert_sql(table: str, row_values: Dict[str, Any]) -> Tuple[str, List[Any]]:
    """
    Builds INSERT ... VALUES ... RETURNING id with parameters.

    Supports:
      - SQL tokens: {"__sql__": "now()"}
      - SQL fragments with params: {"__sql__": "ST_GeomFromText(%s,4326)", "__params__":[...]}
    """
    cols: List[str] = []
    placeholders: List[str] = []
    params: List[Any] = []

    for col, val in row_values.items():
        cols.append(col)
        if isinstance(val, dict) and "__sql__" in val:
            placeholders.append(val["__sql__"])
            frag_params = val.get("__params__")
            if frag_params:
                params.extend(frag_params)
        else:
            placeholders.append("%s")
            params.append(val)

    sql = f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join(placeholders)}) RETURNING id"
    return sql, params


# -----------------------------
# Entity: refs (mapping-driven)
# -----------------------------


def validate_refs_required(excel_row: Dict[str, Any]) -> None:
    r = excel_row
    excel_rownum = r["_excel_row"]
    if normalize_str(r.get("ref_id")) is None:
        raise ColumnIngestionError(
            f"(refs row {excel_rownum}): missing required 'ref_id'."
        )
    date = r.get("date")
    if not isinstance(date, int) and not (isinstance(date, str) and date.isdigit()):
        raise ColumnIngestionError(
            f"(refs row {excel_rownum}): 'date' must be an integer year; got {date!r}"
        )
    if normalize_str(r.get("title")) is None:
        raise ColumnIngestionError(
            f"(refs row {excel_rownum}): 'title' must be non-empty."
        )
    if (
        normalize_str(r.get("authors")) is None
        and normalize_str(r.get("author")) is None
    ):
        raise ColumnIngestionError(
            f"(refs row {excel_rownum}): 'authors' (or 'author') must be non-empty."
        )


def ingest_refs_mapping_driven(
    conn, wb, mapping_entity: Dict[str, Any], sheets: Dict[str, str]
) -> Tuple[Dict[str, int], Dict[str, int]]:
    sheet_name = sheets[mapping_entity["sheet"]]
    ws = wb[sheet_name]
    _, rows = read_sheet_as_dicts(ws)
    if not rows:
        raise ColumnIngestionError(f"'{sheet_name}' tab has no data rows.")

    seen = set()
    for r in rows:
        validate_refs_required(r)
        rid = str(r.get("ref_id")).strip()
        if rid in seen:
            raise ColumnIngestionError(
                f"(refs row {r['_excel_row']}): duplicate 'ref_id' in refs tab: {rid}"
            )
        seen.add(rid)

    table = mapping_entity["table"]
    natural_key = mapping_entity["natural_key"]

    inserted = 0
    reused = 0
    ref_map: Dict[str, int] = {}

    with conn.cursor() as cur:
        for r in rows:
            excel_ref_id = str(r["ref_id"]).strip()

            context = {}
            row_values = eval_entity_row_values(
                mapping_entity, r, context, handlers=None
            )

            where_cols = natural_key
            where_vals = [row_values[c] for c in where_cols]
            where_sql = " AND ".join([f"{c} = %s" for c in where_cols])

            cur.execute(f"SELECT id FROM {table} WHERE {where_sql}", where_vals)
            found = cur.fetchone()
            if found:
                db_id = int(found[0])
                reused += 1
            else:
                sql, params = build_insert_sql(table, row_values)
                cur.execute(sql, params)
                db_id = int(cur.fetchone()[0])
                inserted += 1

            ref_map[excel_ref_id] = db_id

    print(f"refs: processed={len(rows)} inserted={inserted} reused={reused}")
    return ref_map, {"processed": len(rows), "inserted": inserted, "reused": reused}


# -----------------------------
# Entity: col_groups (mapping-driven)
# -----------------------------


def ingest_col_groups_mapping_driven(
    conn, wb, mapping_entity: Dict[str, Any], sheets: Dict[str, str], project_id: int
) -> Tuple[Dict[str, int], Dict[str, int]]:
    sheet_name = sheets[mapping_entity["sheet"]]
    ws = wb[sheet_name]
    _, rows = read_sheet_as_dicts(ws)
    if not rows:
        raise ColumnIngestionError(f"'{sheet_name}' tab has no data rows.")

    distinct_field = mapping_entity["distinct_from_sheet_field"]
    distinct_vals = sorted(
        {
            normalize_str(r.get(distinct_field))
            for r in rows
            if normalize_str(r.get(distinct_field))
        }
    )
    if not distinct_vals:
        raise ColumnIngestionError(
            f"'{sheet_name}' has no non-empty '{distinct_field}' values."
        )

    table = mapping_entity["table"]
    natural_key = mapping_entity["natural_key"]

    out: Dict[str, int] = {}
    inserted = 0
    reused = 0

    with conn.cursor() as cur:
        for g in distinct_vals:
            faux_excel_row = {"col_group": g}

            context = {"project_id": project_id}
            row_values = eval_entity_row_values(
                mapping_entity, faux_excel_row, context, handlers=None
            )

            where_cols = natural_key
            where_vals = [row_values[c] for c in where_cols]
            where_sql = " AND ".join([f"{c} = %s" for c in where_cols])

            cur.execute(f"SELECT id FROM {table} WHERE {where_sql}", where_vals)
            found = cur.fetchone()
            if found:
                db_id = int(found[0])
                reused += 1
            else:
                sql, params = build_insert_sql(table, row_values)
                cur.execute(sql, params)
                db_id = int(cur.fetchone()[0])
                inserted += 1

            out[g] = db_id

    print(
        f"col_groups: processed={len(distinct_vals)} inserted={inserted} reused={reused}"
    )
    return out, {
        "processed": len(distinct_vals),
        "inserted": inserted,
        "reused": reused,
    }


# -----------------------------
# Columns plugin (custom logic + mapping-driven insert for cols)
# -----------------------------


class CleanColRow:
    def __init__(
        self,
        excel_row: int,
        excel_col_id: str,
        col_name: str,
        col_group: str,
        col_type: str,
        ref_excel_ids: List[str],
        lat: Optional[Decimal],
        lng: Optional[Decimal],
        wkt: Optional[str],
    ):
        self.excel_row = excel_row
        self.excel_col_id = excel_col_id
        self.col_name = col_name
        self.col_group = col_group
        self.col_type = col_type
        self.ref_excel_ids = ref_excel_ids
        self.lat = lat
        self.lng = lng
        self.wkt = wkt


def parse_decimal(x: Any, field_name: str, sheet: str, excel_row: int) -> Decimal:
    x = clean_cell_value(x)
    if x is None or (isinstance(x, str) and x.strip() == ""):
        raise ValueError(
            f"({sheet} row {excel_row}): missing required decimal '{field_name}'."
        )
    try:
        return Decimal(str(x).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(
            f"({sheet} row {excel_row}): '{field_name}' must be a decimal; got {x!r}."
        )


def quantize_lat_lng(d: Decimal) -> Decimal:
    # macrostrat.cols.lat and .lng are numeric(8,5), so quantize to 5 decimal places
    # to ensure idempotent equality checks and consistent inserts.
    return d.quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)


def parse_ref_ids(cell: Any, sheet: str, excel_row: int) -> List[str]:
    cell = clean_cell_value(cell)
    if cell is None:
        raise ValueError(f"({sheet} row {excel_row}): missing required 'ref_ids'.")
    if isinstance(cell, (int, float)):
        tokens = [str(int(cell))]
    else:
        tokens = [t.strip() for t in str(cell).split(",")]
    tokens = [t for t in tokens if t]
    if not tokens:
        raise ValueError(
            f"({sheet} row {excel_row}): 'ref_ids' must include one or more comma-separated ids."
        )
    return tokens


def validate_columns_rows(
    headers: List[str],
    rows: List[Dict[str, Any]],
    ref_map: Dict[str, int],
) -> Tuple[List[CleanColRow], bool, Dict[str, Any]]:
    required_cols = {"col_id", "col_name", "col_group", "ref_ids", "col_type"}
    missing = [c for c in sorted(required_cols) if c not in headers]
    if missing:
        raise ColumnIngestionError(
            f"(columns): missing required columns in header row: {missing}"
        )

    has_col_column = "col" in headers
    col_column_values_raw: Dict[str, Any] = {}

    seen_col_ids = set()
    seen_name_group = set()
    seen_latlng = set()
    seen_geom = set()

    cleaned: List[CleanColRow] = []

    for r in rows:
        excel_row = r["_excel_row"]

        col_id_raw = r.get("col_id")
        if col_id_raw is None or (
            isinstance(col_id_raw, str) and col_id_raw.strip() == ""
        ):
            raise ColumnIngestionError(
                f"(columns row {excel_row}): missing required 'col_id'."
            )
        excel_col_id = str(col_id_raw).strip()
        if excel_col_id in seen_col_ids:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): duplicate 'col_id' in columns tab: {excel_col_id}"
            )
        seen_col_ids.add(excel_col_id)

        col_name = normalize_str(r.get("col_name"))
        if not col_name:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): 'col_name' must be non-empty."
            )

        col_group = normalize_str(r.get("col_group"))
        if not col_group:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): 'col_group' must be non-empty."
            )

        ng_key = (col_name, col_group)
        if ng_key in seen_name_group:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): duplicate (col_name, col_group) = {ng_key}"
            )
        seen_name_group.add(ng_key)

        col_type_raw = normalize_str(r.get("col_type"))
        if not col_type_raw:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): 'col_type' must be non-empty."
            )
        col_type = col_type_raw.lower()
        if col_type not in {"column", "section"}:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): 'col_type' must be 'column' or 'section'; got {col_type_raw!r}"
            )

        ref_excel_ids = parse_ref_ids(r.get("ref_ids"), "columns", excel_row)
        missing_refs = [rid for rid in ref_excel_ids if rid not in ref_map]
        if missing_refs:
            raise ColumnIngestionError(
                f"(columns row {excel_row}): ref_ids not found in refs tab: {missing_refs}"
            )

        lat_val = r.get("lat")
        lng_val = r.get("lng")
        geom_val = normalize_str(r.get("geom"))

        has_lat_or_lng = (lat_val is not None and str(lat_val).strip() != "") or (
            lng_val is not None and str(lng_val).strip() != ""
        )
        has_geom = geom_val is not None
        # We allow geom plus optional lat/lng. If geom is present, lat/lng (if provided)
        # will be validated for numeric sanity here and consistency with the geometry later.

        lat: Optional[Decimal] = None
        lng: Optional[Decimal] = None
        wkt: Optional[str] = None

        if has_geom:
            wkt = normalize_wkt(geom_val)
            if wkt in seen_geom:
                raise ColumnIngestionError(
                    f"(columns row {excel_row}): duplicate geom WKT in columns tab."
                )
            seen_geom.add(wkt)
            # If lat/lng are provided along with geom, require both and parse them.
            if has_lat_or_lng:
                lat = quantize_lat_lng(
                    parse_decimal(lat_val, "lat", "columns", excel_row)
                )
                lng = quantize_lat_lng(
                    parse_decimal(lng_val, "lng", "columns", excel_row)
                )

        else:
            # No geom: must have lat/lng
            if not has_lat_or_lng:
                raise ColumnIngestionError(
                    f"(columns row {excel_row}): must provide geom (WKT polygon) or lat/lng."
                )
            lat = quantize_lat_lng(parse_decimal(lat_val, "lat", "columns", excel_row))
            lng = quantize_lat_lng(parse_decimal(lng_val, "lng", "columns", excel_row))
            ll_key = (lat, lng)
            if ll_key in seen_latlng:
                raise ColumnIngestionError(
                    f"(columns row {excel_row}): duplicate (lat,lng)=({lat},{lng}) in columns tab."
                )
            seen_latlng.add(ll_key)
        if has_col_column:
            col_column_values_raw[excel_col_id] = r.get("col")

        cleaned.append(
            CleanColRow(
                excel_row=excel_row,
                excel_col_id=excel_col_id,
                col_name=col_name,
                col_group=col_group,
                col_type=col_type,
                ref_excel_ids=ref_excel_ids,
                lat=lat,
                lng=lng,
                wkt=wkt,
            )
        )

    return cleaned, has_col_column, col_column_values_raw


def validate_wkt_polygon_in_db(cur, wkt: str, sheet_row: int) -> None:
    cur.execute(
        """
        SELECT ST_IsValid(g) AS is_valid, GeometryType(g) AS gtype
        FROM (SELECT ST_GeomFromText(%s, 4326) AS g) s
        """,
        (wkt,),
    )
    is_valid, gtype = cur.fetchone()
    if not is_valid:
        raise ColumnIngestionError(
            f"(columns row {sheet_row}): geom WKT is not a valid geometry."
        )
    if gtype not in ("POLYGON", "MULTIPOLYGON"):
        raise ColumnIngestionError(
            f"(columns row {sheet_row}): geom must be POLYGON or MULTIPOLYGON; got {gtype}."
        )


def validate_geom_matches_latlng_in_db(
    cur, wkt: str, lat: Decimal, lng: Decimal, sheet_row: int
) -> None:
    """
    If user supplies geom + lat/lng, ensure the lat/lng match the point-on-surface of the geom
    at the same precision as macrostrat.cols lat/lng (numeric(8,5)).
    """
    # Compare after rounding to 5 decimal places to match numeric(8,5).
    cur.execute(
        """
        SELECT
          ROUND(ST_Y(ST_PointOnSurface(g))::numeric, 5) AS lat5,
          ROUND(ST_X(ST_PointOnSurface(g))::numeric, 5) AS lng5
        FROM (SELECT ST_GeomFromText(%s, 4326) AS g) s
        """,
        (wkt,),
    )
    lat5, lng5 = cur.fetchone()
    # psycopg2 returns Decimal for numeric
    if lat5 != lat or lng5 != lng:
        raise ColumnIngestionError(
            f"(columns row {sheet_row}): geom + lat/lng are inconsistent.\n"
            f"  Provided lat,lng: {lat},{lng}\n"
            f"  Geom point-on-surface (5dp): {lat5},{lng5}\n"
            "  Fix the spreadsheet so they match (or omit lat/lng when geom is present)."
        )


def compute_cols_col_values(
    cleaned_rows: List[CleanColRow],
    has_col_column: bool,
    col_column_values_raw: Dict[str, Any],
) -> Dict[str, Decimal]:
    out: Dict[str, Decimal] = {}
    if has_col_column:
        for cr in cleaned_rows:
            raw = clean_cell_value(col_column_values_raw.get(cr.excel_col_id))
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                raise ColumnIngestionError(
                    f"(columns row {cr.excel_row}): 'col' column exists but value is empty."
                )
            try:
                d = Decimal(str(raw).strip()).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            except (InvalidOperation, ValueError):
                raise ColumnIngestionError(
                    f"(columns row {cr.excel_row}): 'col' must be numeric; got {raw!r}."
                )
            out[cr.excel_col_id] = d
        return out

    all_ok = True
    tmp: Dict[str, Decimal] = {}
    for cr in cleaned_rows:
        try:
            d = Decimal(str(cr.excel_col_id)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            if abs(d) > Decimal("9999.99"):
                raise InvalidOperation()
            tmp[cr.excel_col_id] = d
        except (InvalidOperation, ValueError):
            all_ok = False
            break
    if all_ok:
        return tmp

    sorted_rows = sorted(
        cleaned_rows,
        key=lambda x: (x.col_group, x.col_name, x.col_type, x.excel_col_id),
    )
    for i, cr in enumerate(sorted_rows, start=1):
        out[cr.excel_col_id] = Decimal(i).quantize(Decimal("0.01"))
    return out


def select_existing_col_id(
    cur,
    project_id: int,
    col_group_id: int,
    col_name: str,
    col_type: str,
    status_code: str,
    lat: Optional[Decimal],
    lng: Optional[Decimal],
    wkt: Optional[str],
) -> Optional[int]:
    if wkt is not None:
        cur.execute(
            """
            SELECT id FROM macrostrat.cols
            WHERE project_id=%s AND col_group_id=%s AND col_name=%s AND col_type=%s AND status_code=%s
              AND poly_geom IS NOT NULL
              AND ST_Equals(poly_geom, ST_GeomFromText(%s, 4326))
            """,
            (project_id, col_group_id, col_name, col_type, status_code, wkt),
        )
    else:
        cur.execute(
            """
            SELECT id FROM macrostrat.cols
            WHERE project_id=%s AND col_group_id=%s AND col_name=%s AND col_type=%s AND status_code=%s
              AND poly_geom IS NULL AND lat=%s AND lng=%s
            """,
            (project_id, col_group_id, col_name, col_type, status_code, lat, lng),
        )
    row = cur.fetchone()
    return int(row[0]) if row else None


def make_cols_handlers(
    col_groups_map: Dict[str, int]
) -> Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]:
    def h_col_group_id(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        cg = normalize_str(excel_row.get("col_group"))
        if not cg:
            raise ColumnIngestionError(
                f"(columns row {excel_row['_excel_row']}): missing 'col_group' for col_group_id resolution."
            )
        if cg not in col_groups_map:
            raise ColumnIngestionError(
                f"(columns row {excel_row['_excel_row']}): col_group '{cg}' not present in resolved col_groups_map."
            )
        return col_groups_map[cg]

    def h_col_value(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        col_value_map = context.get("col_value_map", {})
        excel_col_id = str(excel_row.get("col_id")).strip()
        if excel_col_id not in col_value_map:
            raise ColumnIngestionError(
                f"(columns row {excel_row['_excel_row']}): unable to resolve cols.col value for col_id={excel_col_id}."
            )
        return float(col_value_map[excel_col_id])

    def h_lat(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        lat_val = excel_row.get("lat")
        if wkt:
            wktn = normalize_wkt(wkt)
            return {
                "__sql__": "ST_Y(ST_PointOnSurface(ST_GeomFromText(%s,4326)))",
                "__params__": [wktn],
            }
        return quantize_lat_lng(
            parse_decimal(lat_val, "lat", "columns", excel_row["_excel_row"])
        )

    def h_lng(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        lng_val = excel_row.get("lng")
        if wkt:
            wktn = normalize_wkt(wkt)
            return {
                "__sql__": "ST_X(ST_PointOnSurface(ST_GeomFromText(%s,4326)))",
                "__params__": [wktn],
            }
        return quantize_lat_lng(
            parse_decimal(lng_val, "lng", "columns", excel_row["_excel_row"])
        )

    def h_coordinate(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        if wkt:
            wktn = normalize_wkt(wkt)
            return {
                "__sql__": "ST_SetSRID(ST_MakePoint(ST_X(ST_PointOnSurface(ST_GeomFromText(%s,4326))), ST_Y(ST_PointOnSurface(ST_GeomFromText(%s,4326)))),4326)",
                "__params__": [wktn, wktn],
            }
        lat = quantize_lat_lng(
            parse_decimal(
                excel_row.get("lat"), "lat", "columns", excel_row["_excel_row"]
            )
        )
        lng = quantize_lat_lng(
            parse_decimal(
                excel_row.get("lng"), "lng", "columns", excel_row["_excel_row"]
            )
        )
        return {
            "__sql__": "ST_SetSRID(ST_MakePoint(%s,%s),4326)",
            "__params__": [lng, lat],
        }

    def h_poly_geom(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        if not wkt:
            return None
        wktn = normalize_wkt(wkt)
        return {"__sql__": "ST_GeomFromText(%s,4326)", "__params__": [wktn]}

    def h_area_km2(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        if not wkt:
            return 0.0
        wktn = normalize_wkt(wkt)
        return {
            "__sql__": "(ST_Area(ST_GeomFromText(%s,4326)::geography)/1000000.0)",
            "__params__": [wktn],
        }

    return {
        "cols_col_group_id_v1": h_col_group_id,
        "cols_col_value_v1": h_col_value,
        "cols_lat_v1": h_lat,
        "cols_lng_v1": h_lng,
        "cols_coordinate_v1": h_coordinate,
        "cols_poly_geom_v1": h_poly_geom,
        "cols_area_km2_v1": h_area_km2,
    }


def ingest_columns_plugin(
    conn,
    wb,
    sheet_name: str,
    project_id: int,
    ref_map: Dict[str, int],
    col_groups_map: Dict[str, int],
    cols_entity: Dict[str, Any],
    cols_handlers: Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]],
) -> Tuple[Dict[str, int], Dict[str, Any]]:
    ws = wb[sheet_name]
    headers, rows = read_sheet_as_dicts(ws)
    if not rows:
        raise ColumnIngestionError(f"'{sheet_name}' tab has no data rows.")

    cleaned_rows, has_col_column, col_column_values_raw = validate_columns_rows(
        headers, rows, ref_map
    )
    col_value_map = compute_cols_col_values(
        cleaned_rows, has_col_column, col_column_values_raw
    )

    col_map: Dict[str, int] = {}
    status_code = "in process"

    counts = {
        "processed": len(cleaned_rows),
        "inserted": 0,
        "reused": 0,
        "col_refs_attempted": 0,
        "col_areas_inserted": 0,
        "col_areas_skipped": 0,
    }

    with conn.cursor() as cur:
        for cr in cleaned_rows:
            if cr.wkt is not None:
                validate_wkt_polygon_in_db(cur, cr.wkt, cr.excel_row)
                # If lat/lng were also provided, ensure they match the geometry at numeric(8,5) precision.
                if cr.lat is not None and cr.lng is not None:
                    validate_geom_matches_latlng_in_db(
                        cur, cr.wkt, cr.lat, cr.lng, cr.excel_row
                    )

        for cr in cleaned_rows:
            col_group_id = col_groups_map[cr.col_group]

            existing_id = select_existing_col_id(
                cur,
                project_id,
                col_group_id,
                cr.col_name,
                cr.col_type,
                status_code,
                cr.lat,
                cr.lng,
                cr.wkt,
            )
            if existing_id is not None:
                col_db_id = existing_id
                counts["reused"] += 1
            else:
                faux_excel_row = {
                    "_excel_row": cr.excel_row,
                    "col_id": cr.excel_col_id,
                    "col_name": cr.col_name,
                    "col_group": cr.col_group,
                    "col_type": cr.col_type,
                    "lat": None if cr.wkt is not None else cr.lat,
                    "lng": None if cr.wkt is not None else cr.lng,
                    "geom": cr.wkt,
                }
                context = {"project_id": project_id, "col_value_map": col_value_map}
                row_values = eval_entity_row_values(
                    cols_entity, faux_excel_row, context, handlers=cols_handlers
                )
                sql, params = build_insert_sql(cols_entity["table"], row_values)
                cur.execute(sql, params)
                col_db_id = int(cur.fetchone()[0])
                counts["inserted"] += 1

            col_map[cr.excel_col_id] = col_db_id

            for excel_ref_id in cr.ref_excel_ids:
                ref_db_id = ref_map[excel_ref_id]
                cur.execute(
                    "INSERT INTO macrostrat.col_refs (col_id, ref_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (col_db_id, ref_db_id),
                )
                counts["col_refs_attempted"] += 1

            if cr.wkt is not None:
                cur.execute(
                    "SELECT 1 FROM macrostrat.col_areas WHERE col_id=%s LIMIT 1",
                    (col_db_id,),
                )
                if cur.fetchone() is None:
                    cur.execute(
                        "INSERT INTO macrostrat.col_areas (col_id,wkt,col_area,gmap) VALUES (%s,%s,ST_GeomFromText(%s,4326),%s)",
                        (col_db_id, cr.wkt, cr.wkt, ""),
                    )
                    counts["col_areas_inserted"] += 1
                else:
                    counts["col_areas_skipped"] += 1

    print(
        "columns: "
        f"processed={counts['processed']} "
        f"inserted={counts['inserted']} reused={counts['reused']} "
        f"col_refs_attempted={counts['col_refs_attempted']} "
        f"col_areas_inserted={counts['col_areas_inserted']} col_areas_skipped={counts['col_areas_skipped']}"
    )
    return col_map, counts


# -----------------------------
# Units plugin: dataclasses, lookups, parsing, validation, and ingest helpers
# -----------------------------


@dataclass
class ParsedLithEntry:
    source_kind: str  # "dom" for lithology, "sub" for minor_lith
    raw_entry: str
    lith_id: int
    lith_name: str
    lith_att_ids: List[int]
    lith_att_names: List[str]


@dataclass
class ParsedStratPath:
    raw_entry: str
    names: List[str]  # applied name first, then parents upward


@dataclass
class CleanUnitRow:
    excel_row: int
    excel_unit_id: str
    excel_col_id: str
    excel_section_id: int

    db_col_id: int
    db_b_int_id: int
    db_t_int_id: int

    b_int_name: str
    t_int_name: str
    b_int_age_bottom: Decimal
    b_int_age_top: Decimal
    t_int_age_bottom: Decimal
    t_int_age_top: Decimal

    position_bottom: Decimal
    position_top: Decimal

    b_prop: Optional[Decimal]
    t_prop: Optional[Decimal]

    unit_name: str
    min_thickness: Optional[Decimal]
    max_thickness: Optional[Decimal]

    environments: List[Tuple[int, str]]
    lith_entries: List[ParsedLithEntry]
    minor_lith_entries: List[ParsedLithEntry]
    strat_paths: List[ParsedStratPath]

    notes_text: Optional[str]


def quantize_position(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)


def quantize_thickness(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def normalize_name_key(x: Any) -> Optional[str]:
    s = normalize_str(x)
    return s.lower() if s is not None else None


def parse_optional_decimal(
    x: Any,
    field_name: str,
    sheet: str,
    excel_row: int,
    quantizer: Optional[Callable[[Decimal], Decimal]] = None,
) -> Optional[Decimal]:
    x = clean_cell_value(x)
    if x is None or (isinstance(x, str) and x.strip() == ""):
        return None
    try:
        d = Decimal(str(x).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(
            f"({sheet} row {excel_row}): '{field_name}' must be numeric; got {x!r}."
        )
    return quantizer(d) if quantizer else d


def parse_required_int(x: Any, field_name: str, sheet: str, excel_row: int) -> int:
    x = clean_cell_value(x)
    if x is None or (isinstance(x, str) and x.strip() == ""):
        raise ValueError(
            f"({sheet} row {excel_row}): missing required integer '{field_name}'."
        )
    if isinstance(x, int):
        return x
    if isinstance(x, str) and x.strip().isdigit():
        return int(x.strip())
    raise ValueError(
        f"({sheet} row {excel_row}): '{field_name}' must be an integer; got {x!r}."
    )


def split_semicolon_list(value: Any) -> List[str]:
    s = normalize_str(value)
    if s is None:
        return []
    return [tok.strip() for tok in s.split(";") if tok.strip()]


def split_comma_list(value: str) -> List[str]:
    return [tok.strip() for tok in value.split(",") if tok.strip()]


def parse_lookup_token_id_or_name(
    token: str,
    by_id: Dict[int, Dict[str, Any]],
    by_name: Dict[str, Dict[str, Any]],
    field_name: str,
    sheet: str,
    excel_row: int,
) -> Dict[str, Any]:
    if token.isdigit():
        obj = by_id.get(int(token))
        if obj is None:
            raise ValueError(
                f"({sheet} row {excel_row}): {field_name} token id '{token}' not found in lookup table."
            )
        return obj
    obj = by_name.get(token.lower())
    if obj is None:
        raise ValueError(
            f"({sheet} row {excel_row}): {field_name} token name '{token}' not found in lookup table."
        )
    return obj


def parse_environment_field(
    value: Any, lookups: Dict[str, Any], sheet: str, excel_row: int
) -> List[Tuple[int, str]]:
    out: List[Tuple[int, str]] = []
    seen = set()
    for tok in split_semicolon_list(value):
        obj = parse_lookup_token_id_or_name(
            tok,
            lookups["environs_by_id"],
            lookups["environs_by_name"],
            "environment",
            sheet,
            excel_row,
        )
        key = int(obj["id"])
        if key not in seen:
            out.append((int(obj["id"]), str(obj["name"])))
            seen.add(key)
    return out


def parse_lith_field(
    value: Any,
    source_kind: str,
    lookups: Dict[str, Any],
    sheet: str,
    excel_row: int,
) -> List[ParsedLithEntry]:
    out: List[ParsedLithEntry] = []
    for raw_entry in split_semicolon_list(value):
        toks = split_comma_list(raw_entry)
        if not toks:
            continue
        lith_tok = toks[-1]
        att_toks = toks[:-1]

        lith_obj = parse_lookup_token_id_or_name(
            lith_tok,
            lookups["liths_by_id"],
            lookups["liths_by_name"],
            "lithology",
            sheet,
            excel_row,
        )

        att_ids: List[int] = []
        att_names: List[str] = []
        seen_att_ids = set()
        for att_tok in att_toks:
            att_obj = parse_lookup_token_id_or_name(
                att_tok,
                lookups["lith_atts_by_id"],
                lookups["lith_atts_by_name"],
                "lithology attribute",
                sheet,
                excel_row,
            )
            att_id = int(att_obj["id"])
            if att_id not in seen_att_ids:
                att_ids.append(att_id)
                att_names.append(str(att_obj["name"]))
                seen_att_ids.add(att_id)

        out.append(
            ParsedLithEntry(
                source_kind=source_kind,
                raw_entry=raw_entry,
                lith_id=int(lith_obj["id"]),
                lith_name=str(lith_obj["name"]),
                lith_att_ids=att_ids,
                lith_att_names=att_names,
            )
        )
    return out


def parse_strat_name_field(
    value: Any, sheet: str, excel_row: int
) -> List[ParsedStratPath]:
    out: List[ParsedStratPath] = []
    for raw_entry in split_semicolon_list(value):
        names = split_comma_list(raw_entry)
        if not names:
            continue
        if len(names) > 8:
            raise ValueError(
                f"({sheet} row {excel_row}): strat_name hierarchy path exceeds 8 names: {raw_entry!r}"
            )
        out.append(ParsedStratPath(raw_entry=raw_entry, names=names))
    return out


def build_notes_text(unit_description: Any, comments: Any) -> Optional[str]:
    desc = normalize_str(unit_description)
    comm = normalize_str(comments)
    if desc is None and comm is None:
        return None
    if desc is None:
        return comm
    if comm is None:
        return desc
    return f"{desc}; {comm}"


def resolve_interval(
    token_val: Any,
    field_name: str,
    lookups: Dict[str, Any],
    sheet: str,
    excel_row: int,
) -> Dict[str, Any]:
    tok = normalize_str(token_val)
    if tok is None:
        raise ValueError(f"({sheet} row {excel_row}): missing required '{field_name}'.")
    return parse_lookup_token_id_or_name(
        tok,
        lookups["intervals_by_id"],
        lookups["intervals_by_name"],
        field_name,
        sheet,
        excel_row,
    )


def resolve_unit_positions(
    row: Dict[str, Any], excel_row: int
) -> Tuple[Decimal, Decimal]:
    pos = parse_optional_decimal(
        row.get("position"), "position", "units", excel_row, quantize_position
    )
    t_pos = parse_optional_decimal(
        row.get("t_pos"), "t_pos", "units", excel_row, quantize_position
    )
    b_pos = parse_optional_decimal(
        row.get("b_pos"), "b_pos", "units", excel_row, quantize_position
    )

    if pos is not None:
        return pos, pos

    if t_pos is None or b_pos is None:
        raise ColumnIngestionError(
            f"(units row {excel_row}): must provide either 'position' or both 't_pos' and 'b_pos'."
        )

    if b_pos < t_pos:
        raise ColumnIngestionError(
            f"(units row {excel_row}): require b_pos >= t_pos; got b_pos={b_pos}, t_pos={t_pos}."
        )

    return b_pos, t_pos


def prop_to_smallint(prop: Optional[Decimal]) -> Optional[int]:
    if prop is None:
        return None
    return int((prop * Decimal("10000")).to_integral_value(rounding=ROUND_HALF_UP))


def load_units_lookups(conn) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "intervals_by_id": {},
        "intervals_by_name": {},
        "liths_by_id": {},
        "liths_by_name": {},
        "lith_atts_by_id": {},
        "lith_atts_by_name": {},
        "environs_by_id": {},
        "environs_by_name": {},
    }

    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, interval_name, age_bottom, age_top FROM macrostrat.intervals"
        )
        for iid, interval_name, age_bottom, age_top in cur.fetchall():
            obj = {
                "id": int(iid),
                "name": str(interval_name),
                "age_bottom": Decimal(str(age_bottom)),
                "age_top": Decimal(str(age_top)),
            }
            out["intervals_by_id"][obj["id"]] = obj
            out["intervals_by_name"][str(interval_name).strip().lower()] = obj

        cur.execute("SELECT id, lith FROM macrostrat.liths")
        for lid, lith in cur.fetchall():
            obj = {"id": int(lid), "name": str(lith)}
            out["liths_by_id"][obj["id"]] = obj
            out["liths_by_name"][str(lith).strip().lower()] = obj

        cur.execute("SELECT id, lith_att FROM macrostrat.lith_atts")
        for aid, att in cur.fetchall():
            obj = {"id": int(aid), "name": str(att)}
            out["lith_atts_by_id"][obj["id"]] = obj
            out["lith_atts_by_name"][str(att).strip().lower()] = obj

        cur.execute("SELECT id, environ FROM macrostrat.environs")
        for eid, environ in cur.fetchall():
            obj = {"id": int(eid), "name": str(environ)}
            out["environs_by_id"][obj["id"]] = obj
            out["environs_by_name"][str(environ).strip().lower()] = obj

    return out


def parse_validate_units_rows(
    wb,
    sheet_name: str,
    col_map: Dict[str, int],
    lookups: Dict[str, Any],
) -> List[CleanUnitRow]:
    ws = wb[sheet_name]
    headers, rows = read_sheet_as_dicts(ws)

    required_cols = {
        "unit_id",
        "col_id",
        "section_id",
        "b_int",
        "t_int",
        "unit_name",
        "lithology",
    }
    missing = [c for c in sorted(required_cols) if c not in headers]
    if missing:
        raise ColumnIngestionError(
            f"(units): missing required columns in header row: {missing}"
        )

    if not rows:
        return []

    seen_unit_ids = set()
    seen_unit_identity = set()
    cleaned: List[CleanUnitRow] = []

    for r in rows:
        excel_row = r["_excel_row"]

        excel_unit_id = normalize_str(r.get("unit_id"))
        if excel_unit_id is None:
            raise ColumnIngestionError(
                f"(units row {excel_row}): missing required 'unit_id'."
            )
        if excel_unit_id in seen_unit_ids:
            raise ColumnIngestionError(
                f"(units row {excel_row}): duplicate 'unit_id' in units tab: {excel_unit_id}"
            )
        seen_unit_ids.add(excel_unit_id)

        excel_col_id = normalize_str(r.get("col_id"))
        if excel_col_id is None:
            raise ColumnIngestionError(
                f"(units row {excel_row}): missing required 'col_id'."
            )
        if excel_col_id not in col_map:
            raise ColumnIngestionError(
                f"(units row {excel_row}): col_id '{excel_col_id}' not found in resolved columns map."
            )
        db_col_id = col_map[excel_col_id]

        excel_section_id = parse_required_int(
            r.get("section_id"), "section_id", "units", excel_row
        )

        unit_name = normalize_str(r.get("unit_name"))
        if unit_name is None:
            raise ColumnIngestionError(
                f"(units row {excel_row}): 'unit_name' must be non-empty."
            )

        b_int_obj = resolve_interval(
            r.get("b_int"), "b_int", lookups, "units", excel_row
        )
        t_int_obj = resolve_interval(
            r.get("t_int"), "t_int", lookups, "units", excel_row
        )

        if b_int_obj["age_bottom"] < t_int_obj["age_bottom"]:
            raise ColumnIngestionError(
                f"(units row {excel_row}): age_bottom(b_int) must be >= age_bottom(t_int); "
                f"got b_int={b_int_obj['name']} ({b_int_obj['age_bottom']}), "
                f"t_int={t_int_obj['name']} ({t_int_obj['age_bottom']})."
            )

        position_bottom, position_top = resolve_unit_positions(r, excel_row)

        b_prop = parse_optional_decimal(r.get("b_prop"), "b_prop", "units", excel_row)
        t_prop = parse_optional_decimal(r.get("t_prop"), "t_prop", "units", excel_row)

        if b_prop is not None and not (Decimal("0") <= b_prop <= Decimal("1")):
            raise ColumnIngestionError(
                f"(units row {excel_row}): b_prop must be between 0 and 1; got {b_prop}."
            )
        if t_prop is not None and not (Decimal("0") <= t_prop <= Decimal("1")):
            raise ColumnIngestionError(
                f"(units row {excel_row}): t_prop must be between 0 and 1; got {t_prop}."
            )
        if (
            int(b_int_obj["id"]) == int(t_int_obj["id"])
            and b_prop is not None
            and t_prop is not None
        ):
            if not (b_prop < t_prop):
                raise ColumnIngestionError(
                    f"(units row {excel_row}): when b_int == t_int, require b_prop < t_prop; "
                    f"got b_prop={b_prop}, t_prop={t_prop}."
                )

        min_thickness = parse_optional_decimal(
            r.get("min_thickness"),
            "min_thickness",
            "units",
            excel_row,
            quantize_thickness,
        )
        max_thickness = parse_optional_decimal(
            r.get("max_thickness"),
            "max_thickness",
            "units",
            excel_row,
            quantize_thickness,
        )
        if (
            min_thickness is not None
            and max_thickness is not None
            and max_thickness < min_thickness
        ):
            raise ColumnIngestionError(
                f"(units row {excel_row}): require max_thickness >= min_thickness; "
                f"got max_thickness={max_thickness}, min_thickness={min_thickness}."
            )

        lith_entries = parse_lith_field(
            r.get("lithology"), "dom", lookups, "units", excel_row
        )
        if not lith_entries:
            raise ColumnIngestionError(
                f"(units row {excel_row}): 'lithology' must resolve to one or more entries."
            )

        minor_lith_entries = parse_lith_field(
            r.get("minor_lith"), "sub", lookups, "units", excel_row
        )
        environments = parse_environment_field(
            r.get("environment"), lookups, "units", excel_row
        )
        strat_paths = parse_strat_name_field(r.get("strat_name"), "units", excel_row)
        notes_text = build_notes_text(r.get("unit_description"), r.get("comments"))

        identity_key = (
            db_col_id,
            unit_name,
            position_bottom,
            position_top,
            int(b_int_obj["id"]),
            int(t_int_obj["id"]),
        )
        if identity_key in seen_unit_identity:
            raise ColumnIngestionError(
                f"(units row {excel_row}): duplicate normalized unit identity in units tab: "
                f"{identity_key}"
            )
        seen_unit_identity.add(identity_key)

        cleaned.append(
            CleanUnitRow(
                excel_row=excel_row,
                excel_unit_id=excel_unit_id,
                excel_col_id=excel_col_id,
                excel_section_id=excel_section_id,
                db_col_id=db_col_id,
                db_b_int_id=int(b_int_obj["id"]),
                db_t_int_id=int(t_int_obj["id"]),
                b_int_name=str(b_int_obj["name"]),
                t_int_name=str(t_int_obj["name"]),
                b_int_age_bottom=b_int_obj["age_bottom"],
                b_int_age_top=b_int_obj["age_top"],
                t_int_age_bottom=t_int_obj["age_bottom"],
                t_int_age_top=t_int_obj["age_top"],
                position_bottom=position_bottom,
                position_top=position_top,
                b_prop=b_prop,
                t_prop=t_prop,
                unit_name=unit_name,
                min_thickness=min_thickness,
                max_thickness=max_thickness,
                environments=environments,
                lith_entries=lith_entries,
                minor_lith_entries=minor_lith_entries,
                strat_paths=strat_paths,
                notes_text=notes_text,
            )
        )

    return cleaned


def make_units_handlers() -> Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]:
    def h_units_col_id(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return excel_row["db_col_id"]

    def h_units_fo(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return excel_row["db_b_int_id"]

    def h_units_lo(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return excel_row["db_t_int_id"]

    def h_units_position_bottom(
        excel_row: Dict[str, Any], context: Dict[str, Any]
    ) -> Any:
        return excel_row["position_bottom"]

    def h_units_position_top(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return excel_row["position_top"]

    def h_units_max_thick(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return excel_row.get("max_thickness")

    def h_units_min_thick(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return excel_row.get("min_thickness")

    def h_units_fo_h(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return prop_to_smallint(excel_row.get("b_prop"))

    def h_units_lo_h(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        return prop_to_smallint(excel_row.get("t_prop"))

    return {
        "units_col_id_v1": h_units_col_id,
        "units_fo_v1": h_units_fo,
        "units_lo_v1": h_units_lo,
        "units_position_bottom_v1": h_units_position_bottom,
        "units_position_top_v1": h_units_position_top,
        "units_max_thick_v1": h_units_max_thick,
        "units_min_thick_v1": h_units_min_thick,
        "units_fo_h_v1": h_units_fo_h,
        "units_lo_h_v1": h_units_lo_h,
    }


def ingest_units_main(
    conn,
    cleaned_units: List[CleanUnitRow],
    units_entity: Dict[str, Any],
    units_handlers: Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]],
) -> Tuple[Dict[str, int], Dict[str, int]]:
    unit_map: Dict[str, int] = {}
    inserted = 0
    reused = 0

    table = units_entity["table"]
    natural_key = units_entity["natural_key"]

    with conn.cursor() as cur:
        for cu in cleaned_units:
            faux_excel_row = {
                "_excel_row": cu.excel_row,
                "unit_id": cu.excel_unit_id,
                "col_id": cu.excel_col_id,
                "section_id": cu.excel_section_id,
                "unit_name": cu.unit_name,
                "db_col_id": cu.db_col_id,
                "db_b_int_id": cu.db_b_int_id,
                "db_t_int_id": cu.db_t_int_id,
                "position_bottom": cu.position_bottom,
                "position_top": cu.position_top,
                "min_thickness": cu.min_thickness,
                "max_thickness": cu.max_thickness,
                "b_prop": cu.b_prop,
                "t_prop": cu.t_prop,
            }
            row_values = eval_entity_row_values(
                units_entity, faux_excel_row, {}, handlers=units_handlers
            )

            where_vals = [row_values[c] for c in natural_key]
            where_sql = " AND ".join([f"{c} = %s" for c in natural_key])
            cur.execute(f"SELECT id FROM {table} WHERE {where_sql}", where_vals)
            found = cur.fetchone()
            if found:
                db_id = int(found[0])
                reused += 1
            else:
                sql, params = build_insert_sql(table, row_values)
                cur.execute(sql, params)
                db_id = int(cur.fetchone()[0])
                inserted += 1

            unit_map[cu.excel_unit_id] = db_id

    print(
        f"units: processed={len(cleaned_units)} inserted={inserted} reused={reused} skipped=False"
    )
    return unit_map, {
        "processed": len(cleaned_units),
        "inserted": inserted,
        "reused": reused,
        "skipped": False,
    }


def ingest_unit_liths(
    conn,
    cleaned_units: List[CleanUnitRow],
    unit_map: Dict[str, int],
) -> Tuple[Dict[Tuple[str, str, int], int], Dict[str, int]]:
    """
    Returns:
      unit_lith_map[(excel_unit_id, source_kind, lith_id)] = unit_liths.id
    where source_kind is "dom" or "sub".
    Identity rule:
      (unit_id, lith_id, dom)
    """
    inserted = 0
    reused = 0
    processed = 0
    unit_lith_map: Dict[Tuple[str, str, int], int] = {}

    with conn.cursor() as cur:
        for cu in cleaned_units:
            unit_id = unit_map[cu.excel_unit_id]
            all_entries = list(cu.lith_entries) + list(cu.minor_lith_entries)
            for le in all_entries:
                processed += 1
                dom = le.source_kind
                key = (cu.excel_unit_id, dom, le.lith_id)

                cur.execute(
                    """
                    SELECT id
                    FROM macrostrat.unit_liths
                    WHERE unit_id = %s AND lith_id = %s AND dom = %s
                    """,
                    (unit_id, le.lith_id, dom),
                )
                found = cur.fetchone()
                if found:
                    unit_lith_id = int(found[0])
                    reused += 1
                else:
                    cur.execute(
                        """
                        INSERT INTO macrostrat.unit_liths
                          (lith_id, unit_id, dom, prop, date_mod, comp_prop, mod_prop, toc, ref_id)
                        VALUES
                          (%s, %s, %s, %s, now(), 0, 0, 0, 0)
                        RETURNING id
                        """,
                        (le.lith_id, unit_id, dom, dom),
                    )
                    unit_lith_id = int(cur.fetchone()[0])
                    inserted += 1

                unit_lith_map[key] = unit_lith_id

    print(
        f"unit_liths: {processed} rows processed ({inserted} inserted, {reused} reused)."
    )
    return unit_lith_map, {
        "processed": processed,
        "inserted": inserted,
        "reused": reused,
    }


def ingest_unit_liths_atts(
    conn,
    cleaned_units: List[CleanUnitRow],
    unit_map: Dict[str, int],
    unit_lith_map: Dict[Tuple[str, str, int], int],
) -> Dict[str, int]:
    """
    Identity rule:
      (unit_lith_id, lith_att_id)
    """
    inserted = 0
    reused = 0
    processed = 0

    with conn.cursor() as cur:
        for cu in cleaned_units:
            all_entries = list(cu.lith_entries) + list(cu.minor_lith_entries)
            for le in all_entries:
                unit_lith_id = unit_lith_map[
                    (cu.excel_unit_id, le.source_kind, le.lith_id)
                ]
                for lith_att_id in le.lith_att_ids:
                    processed += 1
                    cur.execute(
                        """
                        SELECT id
                        FROM macrostrat.unit_liths_atts
                        WHERE unit_lith_id = %s AND lith_att_id = %s
                        """,
                        (unit_lith_id, lith_att_id),
                    )
                    found = cur.fetchone()
                    if found:
                        reused += 1
                    else:
                        cur.execute(
                            """
                            INSERT INTO macrostrat.unit_liths_atts
                              (unit_lith_id, lith_att_id, ref_id, date_mod)
                            VALUES
                              (%s, %s, 0, now())
                            RETURNING id
                            """,
                            (unit_lith_id, lith_att_id),
                        )
                        _ = cur.fetchone()[0]
                        inserted += 1

    print(f"unit_liths_atts: processed={processed} inserted={inserted} reused={reused}")
    return {"processed": processed, "inserted": inserted, "reused": reused}


def ingest_unit_environs(
    conn,
    cleaned_units: List[CleanUnitRow],
    unit_map: Dict[str, int],
) -> Dict[str, int]:
    """
    Identity rule:
      (unit_id, environ_id)
    """
    inserted = 0
    reused = 0
    processed = 0

    with conn.cursor() as cur:
        for cu in cleaned_units:
            unit_id = unit_map[cu.excel_unit_id]
            for environ_id, _environ_name in cu.environments:
                processed += 1
                cur.execute(
                    """
                    SELECT id
                    FROM macrostrat.unit_environs
                    WHERE unit_id = %s AND environ_id = %s
                    """,
                    (unit_id, environ_id),
                )
                found = cur.fetchone()
                if found:
                    reused += 1
                else:
                    cur.execute(
                        """
                        INSERT INTO macrostrat.unit_environs
                          (unit_id, environ_id, ref_id, date_mod)
                        VALUES
                          (%s, %s, NULL, now())
                        RETURNING id
                        """,
                        (unit_id, environ_id),
                    )
                    _ = cur.fetchone()[0]
                    inserted += 1

    print(f"unit_environs: processed={processed} inserted={inserted} reused={reused}")
    return {"processed": processed, "inserted": inserted, "reused": reused}


def ingest_unit_notes(
    conn,
    cleaned_units: List[CleanUnitRow],
    unit_map: Dict[str, int],
) -> Dict[str, int]:
    """
    Identity rule:
      (unit_id)
    Operational rule:
      insert if absent; update if present and notes changed; reuse if unchanged.
    """
    inserted = 0
    reused = 0
    updated = 0
    processed = 0

    with conn.cursor() as cur:
        for cu in cleaned_units:
            if cu.notes_text is None:
                continue
            processed += 1
            unit_id = unit_map[cu.excel_unit_id]

            cur.execute(
                """
                SELECT id, notes
                FROM macrostrat.unit_notes
                WHERE unit_id = %s
                """,
                (unit_id,),
            )
            found = cur.fetchone()
            if found is None:
                cur.execute(
                    """
                    INSERT INTO macrostrat.unit_notes
                      (unit_id, notes, date_mod)
                    VALUES
                      (%s, %s, now())
                    RETURNING id
                    """,
                    (unit_id, cu.notes_text),
                )
                _ = cur.fetchone()[0]
                inserted += 1
            else:
                note_id, existing_notes = found
                existing_notes_norm = normalize_str(existing_notes) or ""
                incoming_notes_norm = normalize_str(cu.notes_text) or ""
                if existing_notes_norm == incoming_notes_norm:
                    reused += 1
                else:
                    cur.execute(
                        """
                        UPDATE macrostrat.unit_notes
                        SET notes = %s, date_mod = now()
                        WHERE id = %s
                        """,
                        (cu.notes_text, note_id),
                    )
                    updated += 1

    print(
        f"unit_notes: processed={processed} inserted={inserted} "
        f"reused={reused} updated={updated}"
    )
    return {
        "processed": processed,
        "inserted": inserted,
        "reused": reused,
        "updated": updated,
    }


def rebuild_sections_and_units_sections(
    conn,
    cleaned_units: List[CleanUnitRow],
    unit_map: Dict[str, int],
) -> Tuple[Dict[Tuple[str, int], int], Dict[str, int]]:
    """
    Recompute macrostrat.sections and macrostrat.units_sections for the columns
    represented in the current workbook run.

    Section identity in Excel is:
      (excel_col_id, excel_section_id)

    Section bounds rule:
      fo = oldest b_int among units in the section  -> max(age_bottom of b_int)
      lo = youngest t_int among units in the section -> min(age_top of t_int)
    """
    if not cleaned_units:
        return {}, {
            "affected_cols": 0,
            "units_sections_deleted": 0,
            "sections_deleted": 0,
            "sections_inserted": 0,
            "units_sections_inserted": 0,
            "units_section_id_updated": 0,
        }

    affected_col_ids = sorted({cu.db_col_id for cu in cleaned_units})
    section_groups: Dict[Tuple[str, int], List[CleanUnitRow]] = {}
    for cu in cleaned_units:
        key = (cu.excel_col_id, cu.excel_section_id)
        section_groups.setdefault(key, []).append(cu)

    section_map: Dict[Tuple[str, int], int] = {}
    counts = {
        "affected_cols": len(affected_col_ids),
        "units_sections_deleted": 0,
        "sections_deleted": 0,
        "sections_inserted": 0,
        "units_sections_inserted": 0,
        "units_section_id_updated": 0,
    }

    with conn.cursor() as cur:
        # Delete recomputable join rows first
        cur.execute(
            """
            DELETE FROM macrostrat.units_sections
            WHERE col_id = ANY(%s)
            """,
            (affected_col_ids,),
        )
        counts["units_sections_deleted"] = cur.rowcount

        # Delete existing sections for the affected columns
        cur.execute(
            """
            DELETE FROM macrostrat.sections
            WHERE col_id = ANY(%s)
            """,
            (affected_col_ids,),
        )
        counts["sections_deleted"] = cur.rowcount

        # Rebuild sections
        for (excel_col_id, excel_section_id), rows_in_section in sorted(
            section_groups.items(), key=lambda kv: (kv[0][0], kv[0][1])
        ):
            db_col_id = rows_in_section[0].db_col_id

            # fo = oldest b_int among rows in section => max(b_int_age_bottom)
            fo_row = max(
                rows_in_section, key=lambda r: (r.b_int_age_bottom, r.db_b_int_id)
            )
            fo_id = fo_row.db_b_int_id

            # lo = youngest t_int among rows in section => min(t_int_age_top)
            lo_row = min(
                rows_in_section, key=lambda r: (r.t_int_age_top, r.db_t_int_id)
            )
            lo_id = lo_row.db_t_int_id

            cur.execute(
                """
                INSERT INTO macrostrat.sections (col_id, fo, lo)
                VALUES (%s, %s, %s)
                RETURNING id
                """,
                (db_col_id, fo_id, lo_id),
            )
            db_section_id = int(cur.fetchone()[0])
            section_map[(excel_col_id, excel_section_id)] = db_section_id
            counts["sections_inserted"] += 1

        # Rebuild units_sections
        for cu in cleaned_units:
            db_section_id = section_map[(cu.excel_col_id, cu.excel_section_id)]
            db_unit_id = unit_map[cu.excel_unit_id]
            cur.execute(
                """
                INSERT INTO macrostrat.units_sections (col_id, section_id, unit_id)
                VALUES (%s, %s, %s)
                RETURNING id
                """,
                (cu.db_col_id, db_section_id, db_unit_id),
            )
            _ = cur.fetchone()[0]
            counts["units_sections_inserted"] += 1

        # Update redundant units.section_id
        for cu in cleaned_units:
            db_section_id = section_map[(cu.excel_col_id, cu.excel_section_id)]
            db_unit_id = unit_map[cu.excel_unit_id]
            cur.execute(
                """
                UPDATE macrostrat.units
                SET section_id = %s
                WHERE id = %s
                """,
                (db_section_id, db_unit_id),
            )
            counts["units_section_id_updated"] += cur.rowcount

    print(
        "sections_rebuild: "
        f"affected_cols={counts['affected_cols']} "
        f"units_sections_deleted={counts['units_sections_deleted']} "
        f"sections_deleted={counts['sections_deleted']} "
        f"sections_inserted={counts['sections_inserted']} "
        f"units_sections_inserted={counts['units_sections_inserted']} "
        f"units_section_id_updated={counts['units_section_id_updated']} skipped=False"
    )
    return section_map, counts


def write_strat_name_audit_files(
    audit_dir: str,
    runstamp_compact: str,
    project_id: int,
    cleaned_units: List[CleanUnitRow],
    unit_map: Dict[str, int],
) -> Dict[str, Any]:
    """
    Writes two preliminary strat-name audit TSVs:
      1) strat_name_lexicon_<project_id>_<runstamp>.tsv
         one row per unique normalized hierarchy path
      2) unit_strat_name_map_<project_id>_<runstamp>.tsv
         one row per unit x hierarchy-path assignment
    """
    ensure_dir(audit_dir)

    lexicon_path = os.path.join(
        audit_dir, f"strat_name_lexicon_{project_id}_{runstamp_compact}.tsv"
    )
    unit_map_path = os.path.join(
        audit_dir, f"unit_strat_name_map_{project_id}_{runstamp_compact}.tsv"
    )

    # Aggregate unique normalized paths
    lexicon: Dict[str, Dict[str, Any]] = {}
    assignment_rows: List[Dict[str, Any]] = []

    for cu in cleaned_units:
        db_unit_id = unit_map.get(cu.excel_unit_id)
        for sp in cu.strat_paths:
            names = list(sp.names)
            full_name_path = " | ".join(names)
            applied_name = names[0]
            path_depth = len(names)

            if full_name_path not in lexicon:
                lexicon[full_name_path] = {
                    "applied_name": applied_name,
                    "names": names,
                    "raw_examples": set(),
                    "excel_unit_ids": set(),
                    "excel_col_ids": set(),
                }

            lexicon[full_name_path]["raw_examples"].add(sp.raw_entry)
            lexicon[full_name_path]["excel_unit_ids"].add(cu.excel_unit_id)
            lexicon[full_name_path]["excel_col_ids"].add(cu.excel_col_id)

            assignment_rows.append(
                {
                    "excel_unit_id": cu.excel_unit_id,
                    "db_unit_id": db_unit_id,
                    "excel_col_id": cu.excel_col_id,
                    "applied_name": applied_name,
                    "full_name_path": full_name_path,
                    "path_depth": path_depth,
                    "raw_strat_name_entry": sp.raw_entry,
                    "names": names,
                }
            )

    sorted_paths = sorted(
        lexicon.keys(),
        key=lambda p: (
            lexicon[p]["applied_name"].lower(),
            len(lexicon[p]["names"]),
            p.lower(),
        ),
    )
    path_id_map = {p: f"path_{i:06d}" for i, p in enumerate(sorted_paths, start=1)}

    # Write lexicon TSV
    with open(lexicon_path, "w", encoding="utf-8", newline="") as f:
        headers = [
            "path_id",
            "applied_name",
            "full_name_path",
            "path_depth",
            "raw_examples",
            "unit_count",
            "excel_unit_ids",
            "excel_col_ids",
            "name_1",
            "name_2",
            "name_3",
            "name_4",
            "name_5",
            "name_6",
            "name_7",
            "name_8",
        ]
        f.write("\t".join(headers) + "\n")
        for full_name_path in sorted_paths:
            obj = lexicon[full_name_path]
            names = list(obj["names"])
            exploded = names + [""] * (8 - len(names))
            row = [
                path_id_map[full_name_path],
                obj["applied_name"],
                full_name_path,
                str(len(names)),
                "; ".join(sorted(obj["raw_examples"])),
                str(len(obj["excel_unit_ids"])),
                "; ".join(sorted(obj["excel_unit_ids"])),
                "; ".join(sorted(obj["excel_col_ids"])),
                *exploded[:8],
            ]
            f.write("\t".join(row) + "\n")

    # Write unit -> strat-name map TSV
    with open(unit_map_path, "w", encoding="utf-8", newline="") as f:
        headers = [
            "excel_unit_id",
            "macrostrat_unit_id",
            "excel_col_id",
            "path_id",
            "applied_name",
            "full_name_path",
            "path_depth",
            "raw_strat_name_entry",
            "name_1",
            "name_2",
            "name_3",
            "name_4",
            "name_5",
            "name_6",
            "name_7",
            "name_8",
        ]
        f.write("\t".join(headers) + "\n")
        for row in sorted(
            assignment_rows,
            key=lambda r: (
                r["excel_col_id"],
                r["excel_unit_id"],
                r["full_name_path"].lower(),
            ),
        ):
            names = list(row["names"])
            exploded = names + [""] * (8 - len(names))
            out = [
                row["excel_unit_id"],
                "" if row["db_unit_id"] is None else str(row["db_unit_id"]),
                row["excel_col_id"],
                path_id_map[row["full_name_path"]],
                row["applied_name"],
                row["full_name_path"],
                str(row["path_depth"]),
                row["raw_strat_name_entry"],
                *exploded[:8],
            ]
            f.write("\t".join(out) + "\n")

    print(f"audit: wrote strat-name artifacts to {audit_dir}")
    return {
        "unique_paths": len(sorted_paths),
        "unit_path_assignments": len(assignment_rows),
        "strat_name_lexicon_tsv": os.path.basename(lexicon_path),
        "unit_strat_name_map_tsv": os.path.basename(unit_map_path),
    }


# -----------------------------
# Main
# -----------------------------

from pathlib import Path

from typing import Any
from typer import Argument, Option

__here__ = Path(__file__).parent


def shanan_column_importer(
    excel_file: Path = Argument(None, exists=True),
    *,
    audit_dir: Path = Option(None, exists=False),
    verify_project: bool = Option(False),
    do_audit: bool = Option(True, "--audit/--no-audit"),
):
    """
    Importer to ingest columns based on Shanan's code
    """
    db = get_database()

    # Get the raw Psycopg connection
    conn = connect_db(db.engine)

    try:
        _column_metadata_importer(
            conn,
            excel_file,
            audit_dir=audit_dir,
            verify_project=verify_project,
            do_audit=do_audit,
        )
    except Exception as e:
        conn.rollback()
        raise ColumnIngestionError(
            f"ingestion failed; rolled back transaction. Details: {e}"
        )
    finally:
        conn.close()


def _column_metadata_importer(
    conn,
    excel_file: Path,
    *,
    audit_dir: Optional[Path] = None,
    verify_project: bool = False,
    do_audit: bool = False,
):
    """Column metadata importer main function.
    NOTE: This is not wrapped in transaction handling.
    """

    mapping_filename = str(__here__ / "macrostrat_mapping_v3.json")
    mapping = load_mapping(mapping_filename)
    sheets_map = mapping["sheets"]
    mapping_version = str(mapping["mapping_version"]).strip()

    if excel_file is not None:
        excel_file = str(excel_file)
        excel_sha256 = sha256_file(excel_file)

    if audit_dir is None:
        audit_dir = Path.cwd()
    audit_dir = str(audit_dir)

    try:
        wb = load_workbook(excel_file, data_only=True)
    except FileNotFoundError:
        raise ColumnIngestionError(f"File '{excel_file}' not found.")
    except Exception as e:
        raise ColumnIngestionError(f"Could not read Excel file: {e}")

    meta_sheet = sheets_map["metadata"]
    excel_mapping_version = get_metadata_value(wb, meta_sheet, "mapping_version")
    if excel_mapping_version is None:
        raise ColumnIngestionError(
            f"Excel metadata missing 'mapping_version'. Expected {mapping_version}"
        )
    if str(excel_mapping_version).strip() != mapping_version:
        raise ColumnIngestionError(
            "mapping_version mismatch.\n"
            f"  Excel metadata mapping_version: {str(excel_mapping_version).strip()}\n"
            f"  Mapping file mapping_version:   {mapping_version}"
        )

    project_id_val = get_metadata_value(wb, meta_sheet, "project_id")
    if project_id_val is None:
        raise ColumnIngestionError("Excel metadata missing 'project_id'.")
    if isinstance(project_id_val, int):
        project_id = project_id_val
    elif isinstance(project_id_val, str) and project_id_val.isdigit():
        project_id = int(project_id_val)
    else:
        raise ColumnIngestionError("'project_id' must be an integer.")

    if verify_project:
        verify_project_id_via_api(project_id)

    verify_project_id_in_db(conn, project_id)

    schema = SchemaInspector(conn)
    validate_mapping_against_schema(schema, mapping)

    ents = mapping["entities"]
    refs_ent = ents["refs"]
    col_groups_ent = ents["col_groups"]
    units_ent = ents.get("units")
    cols_ent = ents.get("cols")
    if cols_ent is None:
        raise ColumnIngestionError(
            "Mapping missing required entity 'cols' for Phase 2.5."
        )

    validate_cols_not_null_coverage(schema, cols_ent)
    if units_ent is not None:
        validate_units_not_null_coverage(schema, units_ent)

    # These will be populated inside the committed transaction and then used for audit writing.
    ref_map: Dict[str, int] = {}
    col_map: Dict[str, int] = {}
    unit_map: Dict[str, int] = {}
    refs_counts: Dict[str, int] = {}
    col_groups_counts: Dict[str, int] = {}
    cols_counts: Dict[str, Any] = {}
    units_counts: Dict[str, Any] = {
        "processed": 0,
        "inserted": 0,
        "reused": 0,
        "skipped": True,
    }
    unit_liths_counts: Dict[str, Any] = {"processed": 0, "inserted": 0, "reused": 0}
    unit_liths_atts_counts: Dict[str, Any] = {
        "processed": 0,
        "inserted": 0,
        "reused": 0,
    }
    unit_environs_counts: Dict[str, Any] = {
        "processed": 0,
        "inserted": 0,
        "reused": 0,
    }
    unit_notes_counts: Dict[str, Any] = {
        "processed": 0,
        "inserted": 0,
        "reused": 0,
        "updated": 0,
    }
    sections_rebuild_counts: Dict[str, Any] = {
        "affected_cols": 0,
        "units_sections_deleted": 0,
        "sections_deleted": 0,
        "sections_inserted": 0,
        "units_sections_inserted": 0,
        "units_section_id_updated": 0,
        "skipped": False,
    }
    strat_name_audit_counts: Dict[str, Any] = {
        "unique_paths": 0,
        "unit_path_assignments": 0,
    }
    cleaned_units: List[CleanUnitRow] = []

    ref_map, refs_counts = ingest_refs_mapping_driven(conn, wb, refs_ent, sheets_map)
    col_groups_map, col_groups_counts = ingest_col_groups_mapping_driven(
        conn, wb, col_groups_ent, sheets_map, project_id
    )
    cols_handlers = make_cols_handlers(col_groups_map)

    columns_sheet = sheets_map["columns"]
    col_map, cols_counts = ingest_columns_plugin(
        conn,
        wb,
        columns_sheet,
        project_id,
        ref_map,
        col_groups_map,
        cols_ent,
        cols_handlers,
    )

    if units_ent is not None:
        units_sheet = sheets_map["units"]
        units_lookups = load_units_lookups(conn)
        cleaned_units = parse_validate_units_rows(
            wb=wb,
            sheet_name=units_sheet,
            col_map=col_map,
            lookups=units_lookups,
        )
        if cleaned_units:
            units_handlers = make_units_handlers()
            unit_map, units_counts = ingest_units_main(
                conn, cleaned_units, units_ent, units_handlers
            )
            unit_lith_map, unit_liths_counts = ingest_unit_liths(
                conn, cleaned_units, unit_map
            )
            unit_liths_atts_counts = ingest_unit_liths_atts(
                conn, cleaned_units, unit_map, unit_lith_map
            )
            unit_environs_counts = ingest_unit_environs(conn, cleaned_units, unit_map)
            unit_notes_counts = ingest_unit_notes(conn, cleaned_units, unit_map)
            if cols_counts["inserted"] == 0 and units_counts["inserted"] == 0:
                print(
                    "sections_rebuild: skipping delete/rebuild because "
                    "cols.inserted == 0 and units.inserted == 0."
                )
                sections_rebuild_counts = {
                    "affected_cols": 0,
                    "units_sections_deleted": 0,
                    "sections_deleted": 0,
                    "sections_inserted": 0,
                    "units_sections_inserted": 0,
                    "units_section_id_updated": 0,
                    "skipped": True,
                }
            else:
                _section_map, sections_rebuild_counts = (
                    rebuild_sections_and_units_sections(conn, cleaned_units, unit_map)
                )
        else:
            print("units: processed=0 inserted=0 reused=0 skipped=True")

    print("SUCCESS: Phase 3.0 ingestion completed (transaction committed).")

    if do_audit:
        runstamp_compact, runstamp_iso = utc_runstamp()
        counts_manifest = {
            "refs": refs_counts,
            "col_groups": col_groups_counts,
            "cols": cols_counts,
            "units": units_counts,
            "unit_liths": unit_liths_counts,
            "unit_liths_atts": unit_liths_atts_counts,
            "unit_environs": unit_environs_counts,
            "unit_notes": unit_notes_counts,
            "sections_rebuild": sections_rebuild_counts,
            "strat_name_audit": {"unique_paths": 0, "unit_path_assignments": 0},
        }
    strat_name_lexicon_filename = None
    unit_strat_name_map_filename = None
    strat_name_audit_counts = write_strat_name_audit_files(
        audit_dir=audit_dir,
        runstamp_compact=runstamp_compact,
        project_id=project_id,
        cleaned_units=cleaned_units if units_ent is not None else [],
        unit_map=unit_map,
    )
    strat_name_lexicon_filename = strat_name_audit_counts.get("strat_name_lexicon_tsv")
    unit_strat_name_map_filename = strat_name_audit_counts.get(
        "unit_strat_name_map_tsv"
    )

    # Write manifest so strat-name audit counts are included.
    counts_manifest["strat_name_audit"] = strat_name_audit_counts
    write_audit_artifacts(
        audit_dir=audit_dir,
        runstamp_compact=runstamp_compact,
        runstamp_iso=runstamp_iso,
        script_name=os.path.basename(__file__),
        project_id=project_id,
        mapping_version=mapping_version,
        excel_file=os.path.basename(excel_file),
        excel_sha256=excel_sha256,
        ref_map=ref_map,
        col_map=col_map,
        counts=counts_manifest,
        unit_map=unit_map,
        strat_name_lexicon_filename=strat_name_lexicon_filename,
        unit_strat_name_map_filename=unit_strat_name_map_filename,
    )
    print("SUCCESS: Phase 3 audit artifacts written.")
