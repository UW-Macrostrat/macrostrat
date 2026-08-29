from datetime import date

from openpyxl import load_workbook

from macrostrat.core.database import get_database, set_audit_context
from macrostrat.database import on_conflict

from .age_model import build_age_model
from .columns import (
    assign_section_ids,
    get_column_data,
    reconcile_column_group,
    reconcile_columns,
)
from .database import get_or_create_project
from .metadata import get_metadata
from .refs import get_reference_data, reconcile_references, resolve_column_references
from .units import PositionAxisType, get_units, write_units


def ingest_columns_from_file(
    db,
    data_file,
):
    # Get sheet names
    workbook = load_workbook(
        data_file, read_only=True, data_only=True, keep_links=False
    )
    sheet_names = workbook.sheetnames

    print(f"Sheets: {sheet_names}")

    if "units" not in sheet_names:
        raise ValueError("Sheet 'units' not found in the data file")

    meta = None
    project = None
    if "metadata" in sheet_names:
        meta = get_metadata(data_file)
        project = meta.project

    if "columns" in sheet_names:
        columns = get_column_data(data_file, meta)

    references = []
    if "refs" in sheet_names:
        references = get_reference_data(data_file)

    # Interpret positions as ordinal if the axis type is age
    position = PositionAxisType.HEIGHT
    if meta.axis_type == "age":
        position = PositionAxisType.ORDINAL

    units = get_units(db, data_file, position=position, fill_values=meta.fill_values)

    for col in columns:
        col.units = units.get(col.local_id, [])
        if len(col.units) == 0:
            print(f"Warning: No units found for column {col.local_id}")

    if project is None:
        raise ValueError("Project not found in the data file")

    # One transaction for the whole (column, sections, units) set, so `units.section_id`
    # can reference sections that are created in the same breath — the constraint the
    # legacy importer had to drop because it could not precalculate sections.
    with db.transaction(), on_conflict("restrict"):
        print(f"Ingesting data into project: {project.name}")
        _project = get_or_create_project(db, project)

        # Attribute everything below in the change-tracking trail. Transaction-local
        # is right here (unlike the rebuild scripts): every audited write in this
        # function — col_groups, cols, sections, units, and the unit_boundaries the
        # age model writes — happens inside this one transaction. Set after the
        # project is resolved so the batch can name it; `projects` is not audited,
        # so nothing captured is missed by setting it here rather than earlier.
        set_audit_context(
            db,
            "system:column-ingest",
            f"ingest:{_project.slug}:{date.today().isoformat()}",
        )

        col_group_id = reconcile_column_group(db, _project.id)

        # References come first: columns cite them, and the citations are resolved from
        # workbook-local ids once the reference rows exist.
        ref_map = reconcile_references(db, references)

        reconcile_columns(
            db, columns, project_id=_project.id, col_group_id=col_group_id
        )
        if ref_map:
            resolve_column_references(db, columns, ref_map)

        for col in columns:
            if not col.units:
                continue
            print(f"Ingesting column: {col.name}, ID: {col.id}")
            assign_section_ids(db, col.id, col.units)
            write_units(db, col.units)
            build_age_model(db, col.units)

        db.session.commit()
