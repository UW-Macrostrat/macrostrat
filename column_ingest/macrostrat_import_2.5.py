#!/usr/bin/env python3
"""
Macrostrat ingestion: metadata/project validation + refs + columns
Assumption is that project exists in Macrostrat; project must be defined by valid existing project_id in metadata tab
There is currently no check that supplied project_id is correct.

What this script does (in order):
  1) Reads macrostrat_import.xlsx (tabs: metadata, refs, columns); ALL string values read from Excel are trimmed (leading/trailing whitespace removed)
  immediately upon reading the workbook into row dictionaries.
  2) Extracts project_id from metadata and verifies it via Macrostrat API \\ this step is not needed, added for flair, remove verify_project_id_via_api call to kill
  3) Connects to local Postgres database "macrostrat" and verifies project exists 
  4) In ONE transaction:
       - ingests refs -> macrostrat.refs (idempotent on (pub_year, author, ref))
       - validates columns tab (including uniqueness checks)
       - ingests col_groups -> macrostrat.col_groups (idempotent on (project_id, col_group_long))
       - ingests cols -> macrostrat.cols (idempotent "get-or-create" using:
             project_id, col_group_id, col_name, col_type, status_code, plus geom (polygon) or lat/lng when geom absent
         )
       - ingests col_refs -> macrostrat.col_refs (idempotent DO NOTHING)
       - ingests col_areas -> macrostrat.col_areas (one row per col_id; idempotent via SELECT)

Requires:
- metadata tab of Excel file must have field "mapping_version" matching "mapping_version" in file macrostrat_mapping.json (contains declarative field mappings)

Notes:
- All target tables are schema-qualified under macrostrat.<table>.

Audit:
- Writes four files providing Excel-macrostrat key pairs and a metadata/action summary to ./audit:
  col_id_map_<project_id>_<timestamp>.tsv
  ref_id_map_<project_id>_<timestamp>.tsv
  unit_id_map_<project_id>_<timestamp>.tsv (empty, units not handled in this version)
  macrostrat_import_audit_<project_id>_<timestamp>.json

Postgres connection:
- Hard-coded credentials buried in def connect_db()
	defaults set: dbname="macrostrat", user="postgres", password="", host="localhost", port="5432"

Usage:
  python macrostrat_import_2.5.py --mapping macrostrat_mapping_v2.5.json
"""

import sys
import re
import json
import argparse
import os
import hashlib
from datetime import datetime, timezone
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Dict, List, Optional, Tuple, Callable

import requests
import psycopg2
from openpyxl import load_workbook


# -----------------------------
# Engine utilities
# -----------------------------

def die(msg: str, code: int = 1) -> None:
    print(msg)
    sys.exit(code)


def clean_cell_value(v: Any) -> Any:
    if isinstance(v, str):
        return v.strip()
    return v


def normalize_str(x: Any) -> Optional[str]:
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        return s if s else None
    s = str(x).strip()
    return s if s else None


def normalize_wkt(wkt: str) -> str:
    w = wkt.strip()
    w = re.sub(r"\s+", " ", w)
    return w


def read_sheet_as_dicts(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers: List[str] = []
    for h in rows[0]:
        h = clean_cell_value(h)
        if h is None:
            die(f"ERROR: Sheet '{ws.title}' has empty header cells in first row.")
        headers.append(str(h).strip())

    out: List[Dict[str, Any]] = []
    for i, r in enumerate(rows[1:], start=2):
        if r is None:
            continue
        cleaned_row = [clean_cell_value(v) for v in r]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in cleaned_row):
            continue
        d = {headers[j]: (cleaned_row[j] if j < len(cleaned_row) else None) for j in range(len(headers))}
        d["_excel_row"] = i
        out.append(d)

    return headers, out


def get_metadata_value(wb, sheet_name: str, key: str) -> Any:
    if sheet_name not in wb.sheetnames:
        die(f"ERROR: Workbook does not contain a '{sheet_name}' tab.")
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        k = clean_cell_value(row[0] if len(row) > 0 else None)
        v = clean_cell_value(row[1] if len(row) > 1 else None)
        if k == key:
            return v
    return None


def load_mapping(mapping_path: str) -> Dict[str, Any]:
    try:
        with open(mapping_path, "r", encoding="utf-8") as f:
            mapping = json.load(f)
    except FileNotFoundError:
        die(f"ERROR: Mapping file not found: {mapping_path}")
    except json.JSONDecodeError as e:
        die(f"ERROR: Mapping file is not valid JSON: {e}")

    for k in ("mapping_version", "sheets", "entities"):
        if k not in mapping:
            die(f"ERROR: Mapping file missing required field '{k}'.")
    if not isinstance(mapping["entities"], dict) or not mapping["entities"]:
        die("ERROR: Mapping file 'entities' must be a non-empty object.")

    return mapping

# -----------------------------
# Audit artifacts (Phase 2.5 add-on; Phase 3-ready)
# -----------------------------

def utc_runstamp() -> Tuple[str, str]:
    """
    Returns (runstamp_compact, runstamp_iso_utc)
      - compact: YYYYMMDDTHHMMSSZ
      - iso:     YYYY-MM-DDTHH:MM:SSZ
    """
    dt = datetime.now(timezone.utc).replace(microsecond=0)
    iso = dt.isoformat().replace("00:00", "Z")
    compact = dt.strftime("%Y%m%dT%H%M%SZ")
    return compact, iso


def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def write_tsv_map(path: str, header_a: str, header_b: str, mapping: Dict[str, int]) -> None:
    """
    Writes a 2-column TSV with a header row. Keys sorted for deterministic output.
    """
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(f"{header_a}\t{header_b}\n")
        for k in sorted(mapping.keys()):
            f.write(f"{k}\t{mapping[k]}\n")


def write_units_tsv_header_only(path: str, header_a: str, header_b: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(f"{header_a}\t{header_b}\n")


def write_audit_artifacts(
    audit_dir: str,
    project_id: int,
    mapping_version: str,
    excel_file: str,
    excel_sha256: str,
    ref_map: Dict[str, int],
    col_map: Dict[str, int],
    counts: Dict[str, Any],
    # Phase 3 forward-compat: allow caller to pass a units map later
    unit_map: Optional[Dict[str, int]] = None,
    units_excel_key_name: str = "excel_unit_id",
    units_db_key_name: str = "db_unit_id",
) -> None:
    ensure_dir(audit_dir)
    runstamp_compact, runstamp_iso = utc_runstamp()

    manifest_path = os.path.join(audit_dir, f"macrostrat_import_audit_{project_id}_{runstamp_compact}.json")
    refs_path = os.path.join(audit_dir, f"ref_id_map_{project_id}_{runstamp_compact}.tsv")
    cols_path = os.path.join(audit_dir, f"col_id_map_{project_id}_{runstamp_compact}.tsv")
    units_path = os.path.join(audit_dir, f"unit_id_map_{project_id}_{runstamp_compact}.tsv")

    # TSVs
    write_tsv_map(refs_path, "excel_ref_id", "macrostrat_ref_id", ref_map)
    write_tsv_map(cols_path, "excel_col_id", "macrostrat_col_id", col_map)
    if unit_map is None:
        write_units_tsv_header_only(units_path, units_excel_key_name, units_db_key_name)
    else:
        write_tsv_map(units_path, units_excel_key_name, units_db_key_name, unit_map)

    # Manifest (small)
    manifest = {
        "artifact_version": "1.0",
        "runstamp_utc": runstamp_iso,
        "project_id": project_id,
        "mapping_version": str(mapping_version).strip(),
        "excel_file": excel_file,
        "excel_sha256": excel_sha256,
        "counts": counts,
        "outputs": {
            "manifest_json": os.path.basename(manifest_path),
            "refs_tsv": os.path.basename(refs_path),
            "cols_tsv": os.path.basename(cols_path),
            "units_tsv": os.path.basename(units_path),
        },
    }
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, sort_keys=True)

    print(f"AUDIT: wrote artifacts to {audit_dir}")
    print(f"  {os.path.basename(manifest_path)}")
    print(f"  {os.path.basename(refs_path)}")
    print(f"  {os.path.basename(cols_path)}")
    print(f"  {os.path.basename(units_path)}")

# -----------------------------
# Engine: API + DB
# -----------------------------

def verify_project_id_via_api(project_id: int) -> str:
    api_url = f"https://macrostrat.org/api/defs/projects?project_id={project_id}"
    try:
        r = requests.get(api_url, timeout=10)
        r.raise_for_status()
        j = r.json()
    except Exception as e:
        die(f"ERROR: API request failed: {e}")

    try:
        success_obj = j["success"]
        data_array = success_obj["data"]
        if not data_array:
            die("ERROR: API returned no data for that project_id.")
        returned_project_id = data_array[0]["project_id"]
        returned_project_name = data_array[0]["project"]
    except (KeyError, IndexError, TypeError) as e:
        die(f"ERROR: Unexpected API response structure: {e}")

    if returned_project_id != project_id:
        die(f"ERROR: API returned project_id {returned_project_id}, expected {project_id}.")

    print(f"API OK: project_id {project_id} ('{returned_project_name}')")
    return returned_project_name


def connect_db(dbname="macrostrat", user="postgres", password="", host="localhost", port="5432"):
    try:
        conn = psycopg2.connect(
            dbname=dbname,
            user=user,
            password=password,
            host=host,
            port=port,
        )
        conn.autocommit = False
        return conn
    except psycopg2.Error as e:
        die(f"ERROR: Could not connect to local Postgres '{dbname}': {e}")


def verify_project_id_in_db(conn, project_id: int) -> None:
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM macrostrat.projects WHERE id = %s;", (project_id,))
        if cur.fetchone() is None:
            die(f"ERROR: project_id {project_id} not found in macrostrat.projects.id (local DB).")
    print(f"DB OK: project_id {project_id} exists in macrostrat.projects")


# -----------------------------
# Phase 2.5: schema introspection
# -----------------------------

@dataclass(frozen=True)
class ColumnInfo:
    name: str
    data_type: str
    is_nullable: bool
    has_default: bool
    default_expr: Optional[str]


class SchemaInspector:
    def __init__(self, conn):
        self.conn = conn
        self._cols: Dict[Tuple[str, str], Dict[str, ColumnInfo]] = {}

    @staticmethod
    def split_table(fqtn: str) -> Tuple[str, str]:
        if "." not in fqtn:
            die(f"ERROR: Expected schema-qualified table name, got '{fqtn}'.")
        schema, table = fqtn.split(".", 1)
        return schema, table

    def columns(self, fqtn: str) -> Dict[str, ColumnInfo]:
        schema, table = self.split_table(fqtn)
        key = (schema, table)
        if key in self._cols:
            return self._cols[key]

        with self.conn.cursor() as cur:
            cur.execute(
                """
                SELECT
                  column_name,
                  data_type,
                  (is_nullable = 'YES') AS is_nullable,
                  (column_default IS NOT NULL) AS has_default,
                  column_default
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                ORDER BY ordinal_position
                """,
                (schema, table),
            )
            rows = cur.fetchall()

        if not rows:
            die(f"ERROR: Could not introspect columns for table {schema}.{table} (not found?).")

        cols: Dict[str, ColumnInfo] = {}
        for (name, dtype, is_nullable, has_default, default_expr) in rows:
            cols[name] = ColumnInfo(
                name=name,
                data_type=dtype,
                is_nullable=bool(is_nullable),
                has_default=bool(has_default),
                default_expr=default_expr,
            )

        self._cols[key] = cols
        return cols


def validate_mapping_against_schema(schema: SchemaInspector, mapping: Dict[str, Any]) -> None:
    """
    Validation:
      - all mapped DB columns exist
      - mapped NOT NULL/no-default columns have a source/default in mapping
    """
    for ent_name, ent in mapping["entities"].items():
        fqtn = ent["table"]
        db_cols = schema.columns(fqtn)

        fields = ent.get("fields", {})
        if not isinstance(fields, dict) or not fields:
            die(f"ERROR: Mapping entity '{ent_name}' must define non-empty 'fields'.")

        for db_col in fields.keys():
            if db_col not in db_cols:
                die(f"ERROR: Mapping entity '{ent_name}' refers to missing column '{db_col}' on {fqtn}.")

        for db_col, spec in fields.items():
            cinfo = db_cols[db_col]
            if not cinfo.is_nullable and not cinfo.has_default:
                has_source = any(k in spec for k in ("const", "sql", "expr", "from", "from_any", "handler"))
                has_default = "default" in spec
                if not (has_source or has_default):
                    die(
                        f"ERROR: Entity '{ent_name}' column '{db_col}' is NOT NULL with no default in DB, "
                        f"but mapping provides no source/default."
                    )


def validate_cols_not_null_coverage(schema: SchemaInspector, cols_entity: Dict[str, Any]) -> None:
    """
    Phase 2.5 safety guard:
      For macrostrat.cols specifically, ensure every NOT NULL column with no default
      is covered by the mapping (or is a DB default).
    """
    fqtn = cols_entity["table"]
    db_cols = schema.columns(fqtn)
    mapped = set(cols_entity.get("fields", {}).keys())

    missing_required: List[str] = []
    for name, info in db_cols.items():
        if info.is_nullable:
            continue
        if info.has_default:
            continue
        if name not in mapped:
            missing_required.append(name)

    if missing_required:
        die(
            "ERROR: macrostrat.cols has NOT NULL columns with no defaults that are not mapped.\n"
            f"Missing in mapping: {missing_required}\n"
            "Update macrostrat_mapping.json (entity 'cols') to provide values/defaults."
        )


# -----------------------------
# Mapping functions
# -----------------------------

def concat_ws_dot_space(title: Optional[str], publication: Optional[str]) -> str:
    parts = [p for p in [title, publication] if p is not None and str(p).strip() != ""]
    return ". ".join([str(p).strip() for p in parts])


EXPR_FUNCS = {"concat_ws_dot_space": concat_ws_dot_space}


def apply_transform(value: Any, transform: Optional[str]) -> Any:
    if transform is None:
        return value
    if transform == "lower":
        return value.lower() if isinstance(value, str) else value
    die(f"ERROR: Unknown transform '{transform}'.")


def eval_field_spec(
    spec: Dict[str, Any],
    excel_row: Dict[str, Any],
    context: Dict[str, Any],
    handlers: Optional[Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]] = None,
) -> Any:
    """
    Evaluate a mapping field spec to produce a value (python scalar OR SQL token/fragment dict).

    Supported in Phase 2.5:
      - const: "$project_id" or literal value
      - from: excel column name
      - from_any: [excel_col1, excel_col2, ...] first non-empty
      - expr: function name in EXPR_FUNCS with args (excel fields)
      - sql: SQL token, e.g. "now()"
      - handler: name resolved in handlers dict
      - transform: currently supports "lower"
      - default: used if value is None/empty string
    """
    value = None

    if "handler" in spec:
        if handlers is None:
            die("ERROR: handler specified but no handlers registry provided.")
        hname = spec["handler"]
        hfn = handlers.get(hname)
        if hfn is None:
            die(f"ERROR: Unknown handler '{hname}'.")
        value = hfn(excel_row, context)

    elif "const" in spec:
        c = spec["const"]
        if isinstance(c, str) and c.startswith("$"):
            key = c[1:]
            value = context.get(key)
        else:
            value = c

    elif "from" in spec:
        value = excel_row.get(spec["from"])

    elif "from_any" in spec:
        for k in spec["from_any"]:
            v = excel_row.get(k)
            if v is not None and not (isinstance(v, str) and v.strip() == ""):
                value = v
                break

    elif "expr" in spec:
        fn_name = spec["expr"]
        fn = EXPR_FUNCS.get(fn_name)
        if fn is None:
            die(f"ERROR: Unknown expr function '{fn_name}'.")
        args = []
        for a in spec.get("args", []):
            args.append(excel_row.get(a))
        value = fn(*args)

    elif "sql" in spec:
        value = {"__sql__": spec["sql"]}

    else:
        die(f"ERROR: Unsupported field spec: {spec}")

    value = apply_transform(value, spec.get("transform"))

    if value is None or (isinstance(value, str) and value.strip() == ""):
        if "default" in spec:
            value = spec["default"]

    if isinstance(value, str):
        value = value.strip()

    return value


def eval_entity_row_values(
    entity: Dict[str, Any],
    excel_row: Dict[str, Any],
    context: Dict[str, Any],
    handlers: Optional[Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]] = None,
) -> Dict[str, Any]:
    fields = entity["fields"]
    row_values: Dict[str, Any] = {}
    for db_col, spec in fields.items():
        val = eval_field_spec(spec, excel_row, context, handlers=handlers)
        if spec.get("omit_if_none") and val is None:
            continue
        row_values[db_col] = val
    return row_values


def build_insert_sql(table: str, row_values: Dict[str, Any]) -> Tuple[str, List[Any]]:
    """
    Builds INSERT ... VALUES ... RETURNING id with parameters.

    Supports:
      - SQL tokens: {"__sql__": "now()"}
      - SQL fragments with params: {"__sql__": "ST_GeomFromText(%s,4326)", "__params__":[...]}
    """
    cols: List[str] = []
    placeholders: List[str] = []
    params: List[Any] = []

    for col, val in row_values.items():
        cols.append(col)
        if isinstance(val, dict) and "__sql__" in val:
            placeholders.append(val["__sql__"])
            frag_params = val.get("__params__")
            if frag_params:
                params.extend(frag_params)
        else:
            placeholders.append("%s")
            params.append(val)

    sql = f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join(placeholders)}) RETURNING id"
    return sql, params


# -----------------------------
# Entity: refs (mapping-driven)
# -----------------------------

def validate_refs_required(excel_row: Dict[str, Any]) -> None:
    r = excel_row
    excel_rownum = r["_excel_row"]
    if normalize_str(r.get("ref_id")) is None:
        die(f"ERROR (refs row {excel_rownum}): missing required 'ref_id'.")
    date = r.get("date")
    if not isinstance(date, int) and not (isinstance(date, str) and date.isdigit()):
        die(f"ERROR (refs row {excel_rownum}): 'date' must be an integer year; got {date!r}")
    if normalize_str(r.get("title")) is None:
        die(f"ERROR (refs row {excel_rownum}): 'title' must be non-empty.")
    if normalize_str(r.get("authors")) is None and normalize_str(r.get("author")) is None:
        die(f"ERROR (refs row {excel_rownum}): 'authors' (or 'author') must be non-empty.")


def ingest_refs_mapping_driven(conn, wb, mapping_entity: Dict[str, Any], sheets: Dict[str, str]) -> Tuple[Dict[str, int], Dict[str, int]]:
    sheet_name = sheets[mapping_entity["sheet"]]
    ws = wb[sheet_name]
    _, rows = read_sheet_as_dicts(ws)
    if not rows:
        die(f"ERROR: '{sheet_name}' tab has no data rows.")

    seen = set()
    for r in rows:
        validate_refs_required(r)
        rid = str(r.get("ref_id")).strip()
        if rid in seen:
            die(f"ERROR (refs row {r['_excel_row']}): duplicate 'ref_id' in refs tab: {rid}")
        seen.add(rid)

    table = mapping_entity["table"]
    natural_key = mapping_entity["natural_key"]

    inserted = 0
    reused = 0
    ref_map: Dict[str, int] = {}

    with conn.cursor() as cur:
        for r in rows:
            excel_ref_id = str(r["ref_id"]).strip()

            context = {}
            row_values = eval_entity_row_values(mapping_entity, r, context, handlers=None)

            where_cols = natural_key
            where_vals = [row_values[c] for c in where_cols]
            where_sql = " AND ".join([f"{c} = %s" for c in where_cols])

            cur.execute(f"SELECT id FROM {table} WHERE {where_sql}", where_vals)
            found = cur.fetchone()
            if found:
                db_id = int(found[0])
                reused += 1
            else:
                sql, params = build_insert_sql(table, row_values)
                cur.execute(sql, params)
                db_id = int(cur.fetchone()[0])
                inserted += 1

            ref_map[excel_ref_id] = db_id

    print(f"refs: {len(rows)} rows processed ({inserted} inserted, {reused} reused).")
    return ref_map, {"processed": len(rows), "inserted": inserted, "reused": reused}


# -----------------------------
# Entity: col_groups (mapping-driven)
# -----------------------------

def ingest_col_groups_mapping_driven(conn, wb, mapping_entity: Dict[str, Any], sheets: Dict[str, str], project_id: int) -> Dict[str, int]:
    """
    Returns mapping col_group_long (string) -> col_groups.id
    Uses SELECT/INSERT idempotently (no unique constraints assumed).
    """
def ingest_col_groups_mapping_driven(conn, wb, mapping_entity: Dict[str, Any], sheets: Dict[str, str], project_id: int) -> Tuple[Dict[str, int], Dict[str, int]]:
    sheet_name = sheets[mapping_entity["sheet"]]
    ws = wb[sheet_name]
    _, rows = read_sheet_as_dicts(ws)
    if not rows:
        die(f"ERROR: '{sheet_name}' tab has no data rows.")

    distinct_field = mapping_entity["distinct_from_sheet_field"]
    distinct_vals = sorted({normalize_str(r.get(distinct_field)) for r in rows if normalize_str(r.get(distinct_field))})
    if not distinct_vals:
        die(f"ERROR: '{sheet_name}' has no non-empty '{distinct_field}' values.")

    table = mapping_entity["table"]
    natural_key = mapping_entity["natural_key"]

    out: Dict[str, int] = {}
    inserted = 0
    reused = 0

    with conn.cursor() as cur:
        for g in distinct_vals:
            faux_excel_row = {"col_group": g}

            context = {"project_id": project_id}
            row_values = eval_entity_row_values(mapping_entity, faux_excel_row, context, handlers=None)

            where_cols = natural_key
            where_vals = [row_values[c] for c in where_cols]
            where_sql = " AND ".join([f"{c} = %s" for c in where_cols])

            cur.execute(f"SELECT id FROM {table} WHERE {where_sql}", where_vals)
            found = cur.fetchone()
            if found:
                db_id = int(found[0])
                reused += 1
            else:
                sql, params = build_insert_sql(table, row_values)
                cur.execute(sql, params)
                db_id = int(cur.fetchone()[0])
                inserted += 1

            out[g] = db_id

    print(f"col_groups: {len(distinct_vals)} groups ({inserted} inserted, {reused} reused).")
    return out, {"processed": len(distinct_vals), "inserted": inserted, "reused": reused}


# -----------------------------
# Columns plugin (custom logic + mapping-driven insert for cols)
# -----------------------------

class CleanColRow:
    def __init__(
        self,
        excel_row: int,
        excel_col_id: str,
        col_name: str,
        col_group: str,
        col_type: str,
        ref_excel_ids: List[str],
        lat: Optional[Decimal],
        lng: Optional[Decimal],
        wkt: Optional[str],
    ):
        self.excel_row = excel_row
        self.excel_col_id = excel_col_id
        self.col_name = col_name
        self.col_group = col_group
        self.col_type = col_type
        self.ref_excel_ids = ref_excel_ids
        self.lat = lat
        self.lng = lng
        self.wkt = wkt


def parse_decimal(x: Any, field_name: str, sheet: str, excel_row: int) -> Decimal:
    x = clean_cell_value(x)
    if x is None or (isinstance(x, str) and x.strip() == ""):
        raise ValueError(f"ERROR ({sheet} row {excel_row}): missing required decimal '{field_name}'.")
    try:
        return Decimal(str(x).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"ERROR ({sheet} row {excel_row}): '{field_name}' must be a decimal; got {x!r}.")


def quantize_lat_lng(d: Decimal) -> Decimal:
    # macrostrat.cols.lat and .lng are numeric(8,5), so quantize to 5 decimal places
    # to ensure idempotent equality checks and consistent inserts.
    return d.quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)


def parse_ref_ids(cell: Any, sheet: str, excel_row: int) -> List[str]:
    cell = clean_cell_value(cell)
    if cell is None:
        raise ValueError(f"ERROR ({sheet} row {excel_row}): missing required 'ref_ids'.")
    if isinstance(cell, (int, float)):
        tokens = [str(int(cell))]
    else:
        tokens = [t.strip() for t in str(cell).split(",")]
    tokens = [t for t in tokens if t]
    if not tokens:
        raise ValueError(f"ERROR ({sheet} row {excel_row}): 'ref_ids' must include one or more comma-separated ids.")
    return tokens


def validate_columns_rows(
    headers: List[str],
    rows: List[Dict[str, Any]],
    ref_map: Dict[str, int],
) -> Tuple[List[CleanColRow], bool, Dict[str, Any]]:
    required_cols = {"col_id", "col_name", "col_group", "ref_ids", "col_type"}
    missing = [c for c in sorted(required_cols) if c not in headers]
    if missing:
        die(f"ERROR (columns): missing required columns in header row: {missing}")

    has_col_column = "col" in headers
    col_column_values_raw: Dict[str, Any] = {}

    seen_col_ids = set()
    seen_name_group = set()
    seen_latlng = set()
    seen_geom = set()

    cleaned: List[CleanColRow] = []

    for r in rows:
        excel_row = r["_excel_row"]

        col_id_raw = r.get("col_id")
        if col_id_raw is None or (isinstance(col_id_raw, str) and col_id_raw.strip() == ""):
            die(f"ERROR (columns row {excel_row}): missing required 'col_id'.")
        excel_col_id = str(col_id_raw).strip()
        if excel_col_id in seen_col_ids:
            die(f"ERROR (columns row {excel_row}): duplicate 'col_id' in columns tab: {excel_col_id}")
        seen_col_ids.add(excel_col_id)

        col_name = normalize_str(r.get("col_name"))
        if not col_name:
            die(f"ERROR (columns row {excel_row}): 'col_name' must be non-empty.")

        col_group = normalize_str(r.get("col_group"))
        if not col_group:
            die(f"ERROR (columns row {excel_row}): 'col_group' must be non-empty.")

        ng_key = (col_name, col_group)
        if ng_key in seen_name_group:
            die(f"ERROR (columns row {excel_row}): duplicate (col_name, col_group) = {ng_key}")
        seen_name_group.add(ng_key)

        col_type_raw = normalize_str(r.get("col_type"))
        if not col_type_raw:
            die(f"ERROR (columns row {excel_row}): 'col_type' must be non-empty.")
        col_type = col_type_raw.lower()
        if col_type not in {"column", "section"}:
            die(f"ERROR (columns row {excel_row}): 'col_type' must be 'column' or 'section'; got {col_type_raw!r}")

        ref_excel_ids = parse_ref_ids(r.get("ref_ids"), "columns", excel_row)
        missing_refs = [rid for rid in ref_excel_ids if rid not in ref_map]
        if missing_refs:
            die(f"ERROR (columns row {excel_row}): ref_ids not found in refs tab: {missing_refs}")

        lat_val = r.get("lat")
        lng_val = r.get("lng")
        geom_val = normalize_str(r.get("geom"))

        has_lat_or_lng = (lat_val is not None and str(lat_val).strip() != "") or (lng_val is not None and str(lng_val).strip() != "")
        has_geom = geom_val is not None
        # We allow geom plus optional lat/lng. If geom is present, lat/lng (if provided)
        # will be validated for numeric sanity here and consistency with the geometry later.

        lat: Optional[Decimal] = None
        lng: Optional[Decimal] = None
        wkt: Optional[str] = None

        if has_geom:
            wkt = normalize_wkt(geom_val)
            if wkt in seen_geom:
                die(f"ERROR (columns row {excel_row}): duplicate geom WKT in columns tab.")
            seen_geom.add(wkt)
            # If lat/lng are provided along with geom, require both and parse them.
            if has_lat_or_lng:
                lat = quantize_lat_lng(parse_decimal(lat_val, "lat", "columns", excel_row))
                lng = quantize_lat_lng(parse_decimal(lng_val, "lng", "columns", excel_row))

        else:
            # No geom: must have lat/lng
            if not has_lat_or_lng:
                die(f"ERROR (columns row {excel_row}): must provide geom (WKT polygon) or lat/lng.")
            lat = quantize_lat_lng(parse_decimal(lat_val, "lat", "columns", excel_row))
            lng = quantize_lat_lng(parse_decimal(lng_val, "lng", "columns", excel_row))
            ll_key = (lat, lng)
            if ll_key in seen_latlng:
                die(f"ERROR (columns row {excel_row}): duplicate (lat,lng)=({lat},{lng}) in columns tab.")
            seen_latlng.add(ll_key)
        if has_col_column:
            col_column_values_raw[excel_col_id] = r.get("col")

        cleaned.append(
            CleanColRow(
                excel_row=excel_row,
                excel_col_id=excel_col_id,
                col_name=col_name,
                col_group=col_group,
                col_type=col_type,
                ref_excel_ids=ref_excel_ids,
                lat=lat,
                lng=lng,
                wkt=wkt,
            )
        )

    return cleaned, has_col_column, col_column_values_raw


def validate_wkt_polygon_in_db(cur, wkt: str, sheet_row: int) -> None:
    cur.execute(
        """
        SELECT ST_IsValid(g) AS is_valid, GeometryType(g) AS gtype
        FROM (SELECT ST_GeomFromText(%s, 4326) AS g) s
        """,
        (wkt,),
    )
    is_valid, gtype = cur.fetchone()
    if not is_valid:
        die(f"ERROR (columns row {sheet_row}): geom WKT is not a valid geometry.")
    if gtype not in ("POLYGON", "MULTIPOLYGON"):
        die(f"ERROR (columns row {sheet_row}): geom must be POLYGON or MULTIPOLYGON; got {gtype}.")

def validate_geom_matches_latlng_in_db(cur, wkt: str, lat: Decimal, lng: Decimal, sheet_row: int) -> None:
    """
    If user supplies geom + lat/lng, ensure the lat/lng match the point-on-surface of the geom
    at the same precision as macrostrat.cols lat/lng (numeric(8,5)).
    """
    # Compare after rounding to 5 decimal places to match numeric(8,5).
    cur.execute(
        """
        SELECT
          ROUND(ST_Y(ST_PointOnSurface(g))::numeric, 5) AS lat5,
          ROUND(ST_X(ST_PointOnSurface(g))::numeric, 5) AS lng5
        FROM (SELECT ST_GeomFromText(%s, 4326) AS g) s
        """,
        (wkt,),
    )
    lat5, lng5 = cur.fetchone()
    # psycopg2 returns Decimal for numeric
    if lat5 != lat or lng5 != lng:
        die(
            f"ERROR (columns row {sheet_row}): geom + lat/lng are inconsistent.\n"
            f"  Provided lat,lng: {lat},{lng}\n"
            f"  Geom point-on-surface (5dp): {lat5},{lng5}\n"
            "  Fix the spreadsheet so they match (or omit lat/lng when geom is present)."
        )

def compute_cols_col_values(cleaned_rows: List[CleanColRow], has_col_column: bool, col_column_values_raw: Dict[str, Any]) -> Dict[str, Decimal]:
    out: Dict[str, Decimal] = {}
    if has_col_column:
        for cr in cleaned_rows:
            raw = clean_cell_value(col_column_values_raw.get(cr.excel_col_id))
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                die(f"ERROR (columns row {cr.excel_row}): 'col' column exists but value is empty.")
            try:
                d = Decimal(str(raw).strip()).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            except (InvalidOperation, ValueError):
                die(f"ERROR (columns row {cr.excel_row}): 'col' must be numeric; got {raw!r}.")
            out[cr.excel_col_id] = d
        return out

    all_ok = True
    tmp: Dict[str, Decimal] = {}
    for cr in cleaned_rows:
        try:
            d = Decimal(str(cr.excel_col_id)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if abs(d) > Decimal("9999.99"):
                raise InvalidOperation()
            tmp[cr.excel_col_id] = d
        except (InvalidOperation, ValueError):
            all_ok = False
            break
    if all_ok:
        return tmp

    sorted_rows = sorted(cleaned_rows, key=lambda x: (x.col_group, x.col_name, x.col_type, x.excel_col_id))
    for i, cr in enumerate(sorted_rows, start=1):
        out[cr.excel_col_id] = Decimal(i).quantize(Decimal("0.01"))
    return out


def select_existing_col_id(cur, project_id: int, col_group_id: int, col_name: str, col_type: str, status_code: str,
                          lat: Optional[Decimal], lng: Optional[Decimal], wkt: Optional[str]) -> Optional[int]:
    if wkt is not None:
        cur.execute(
            """
            SELECT id FROM macrostrat.cols
            WHERE project_id=%s AND col_group_id=%s AND col_name=%s AND col_type=%s AND status_code=%s
              AND poly_geom IS NOT NULL
              AND ST_Equals(poly_geom, ST_GeomFromText(%s, 4326))
            """,
            (project_id, col_group_id, col_name, col_type, status_code, wkt),
        )
    else:
        cur.execute(
            """
            SELECT id FROM macrostrat.cols
            WHERE project_id=%s AND col_group_id=%s AND col_name=%s AND col_type=%s AND status_code=%s
              AND poly_geom IS NULL AND lat=%s AND lng=%s
            """,
            (project_id, col_group_id, col_name, col_type, status_code, lat, lng),
        )
    row = cur.fetchone()
    return int(row[0]) if row else None


def make_cols_handlers(col_groups_map: Dict[str, int]) -> Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]]:
    def h_col_group_id(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        cg = normalize_str(excel_row.get("col_group"))
        if not cg:
            die(f"ERROR (columns row {excel_row['_excel_row']}): missing 'col_group' for col_group_id resolution.")
        if cg not in col_groups_map:
            die(f"ERROR (columns row {excel_row['_excel_row']}): col_group '{cg}' not present in resolved col_groups_map.")
        return col_groups_map[cg]

    def h_col_value(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        col_value_map = context.get("col_value_map", {})
        excel_col_id = str(excel_row.get("col_id")).strip()
        if excel_col_id not in col_value_map:
            die(f"ERROR (columns row {excel_row['_excel_row']}): unable to resolve cols.col value for col_id={excel_col_id}.")
        return float(col_value_map[excel_col_id])

    def h_lat(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        lat_val = excel_row.get("lat")
        if wkt:
            wktn = normalize_wkt(wkt)
            return {"__sql__": "ST_Y(ST_PointOnSurface(ST_GeomFromText(%s,4326)))", "__params__": [wktn]}
        return quantize_lat_lng(parse_decimal(lat_val, "lat", "columns", excel_row["_excel_row"]))

    def h_lng(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        lng_val = excel_row.get("lng")
        if wkt:
            wktn = normalize_wkt(wkt)
            return {"__sql__": "ST_X(ST_PointOnSurface(ST_GeomFromText(%s,4326)))", "__params__": [wktn]}
        return quantize_lat_lng(parse_decimal(lng_val, "lng", "columns", excel_row["_excel_row"]))

    def h_coordinate(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        if wkt:
            wktn = normalize_wkt(wkt)
            return {
                "__sql__": "ST_SetSRID(ST_MakePoint(ST_X(ST_PointOnSurface(ST_GeomFromText(%s,4326))), ST_Y(ST_PointOnSurface(ST_GeomFromText(%s,4326)))),4326)",
                "__params__": [wktn, wktn],
            }
        lat = quantize_lat_lng(parse_decimal(excel_row.get("lat"), "lat", "columns", excel_row["_excel_row"]))
        lng = quantize_lat_lng(parse_decimal(excel_row.get("lng"), "lng", "columns", excel_row["_excel_row"]))
        return {"__sql__": "ST_SetSRID(ST_MakePoint(%s,%s),4326)", "__params__": [lng, lat]}

    def h_poly_geom(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        if not wkt:
            return None
        wktn = normalize_wkt(wkt)
        return {"__sql__": "ST_GeomFromText(%s,4326)", "__params__": [wktn]}

    def h_area_km2(excel_row: Dict[str, Any], context: Dict[str, Any]) -> Any:
        wkt = normalize_str(excel_row.get("geom"))
        if not wkt:
            return 0.0
        wktn = normalize_wkt(wkt)
        return {"__sql__": "(ST_Area(ST_GeomFromText(%s,4326)::geography)/1000000.0)", "__params__": [wktn]}

    return {
        "cols_col_group_id_v1": h_col_group_id,
        "cols_col_value_v1": h_col_value,
        "cols_lat_v1": h_lat,
        "cols_lng_v1": h_lng,
        "cols_coordinate_v1": h_coordinate,
        "cols_poly_geom_v1": h_poly_geom,
        "cols_area_km2_v1": h_area_km2,
    }


def ingest_columns_plugin(
    conn,
    wb,
    sheet_name: str,
    project_id: int,
    ref_map: Dict[str, int],
    col_groups_map: Dict[str, int],
    cols_entity: Dict[str, Any],
    cols_handlers: Dict[str, Callable[[Dict[str, Any], Dict[str, Any]], Any]],
) -> Tuple[Dict[str, int], Dict[str, Any]]:
    ws = wb[sheet_name]
    headers, rows = read_sheet_as_dicts(ws)
    if not rows:
        die(f"ERROR: '{sheet_name}' tab has no data rows.")

    cleaned_rows, has_col_column, col_column_values_raw = validate_columns_rows(headers, rows, ref_map)
    col_value_map = compute_cols_col_values(cleaned_rows, has_col_column, col_column_values_raw)

    col_map: Dict[str, int] = {}
    status_code = "in process"

    counts = {"processed": len(cleaned_rows), "inserted": 0, "reused": 0, "col_refs_attempted": 0, "col_areas_inserted": 0, "col_areas_skipped": 0}

    with conn.cursor() as cur:
        for cr in cleaned_rows:
            if cr.wkt is not None:
                validate_wkt_polygon_in_db(cur, cr.wkt, cr.excel_row)
                # If lat/lng were also provided, ensure they match the geometry at numeric(8,5) precision.
                if cr.lat is not None and cr.lng is not None:
                    validate_geom_matches_latlng_in_db(cur, cr.wkt, cr.lat, cr.lng, cr.excel_row)

        for cr in cleaned_rows:
            col_group_id = col_groups_map[cr.col_group]

            existing_id = select_existing_col_id(
                cur, project_id, col_group_id, cr.col_name, cr.col_type, status_code, cr.lat, cr.lng, cr.wkt
            )
            if existing_id is not None:
                col_db_id = existing_id
                counts["reused"] += 1
            else:
                faux_excel_row = {
                    "_excel_row": cr.excel_row,
                    "col_id": cr.excel_col_id,
                    "col_name": cr.col_name,
                    "col_group": cr.col_group,
                    "col_type": cr.col_type,
                    "lat": None if cr.wkt is not None else cr.lat,
                    "lng": None if cr.wkt is not None else cr.lng,
                    "geom": cr.wkt,
                }
                context = {"project_id": project_id, "col_value_map": col_value_map}
                row_values = eval_entity_row_values(cols_entity, faux_excel_row, context, handlers=cols_handlers)
                sql, params = build_insert_sql(cols_entity["table"], row_values)
                cur.execute(sql, params)
                col_db_id = int(cur.fetchone()[0])
                counts["inserted"] += 1

            col_map[cr.excel_col_id] = col_db_id

            for excel_ref_id in cr.ref_excel_ids:
                ref_db_id = ref_map[excel_ref_id]
                cur.execute(
                    "INSERT INTO macrostrat.col_refs (col_id, ref_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (col_db_id, ref_db_id),
                )
                counts["col_refs_attempted"] += 1

            if cr.wkt is not None:
                cur.execute("SELECT 1 FROM macrostrat.col_areas WHERE col_id=%s LIMIT 1", (col_db_id,))
                if cur.fetchone() is None:
                    cur.execute(
                        "INSERT INTO macrostrat.col_areas (col_id,wkt,col_area,gmap) VALUES (%s,%s,ST_GeomFromText(%s,4326),%s)",
                        (col_db_id, cr.wkt, cr.wkt, ""),
                    )
                    counts["col_areas_inserted"] += 1
                else:
                    counts["col_areas_skipped"] += 1

    print(
        "columns:\n"
        f"  cols:      {counts['inserted']} inserted, {counts['reused']} reused\n"
        f"  col_refs:  {counts['col_refs_attempted']} link inserts attempted\n"
        f"  col_areas: {counts['col_areas_inserted']} inserted, {counts['col_areas_skipped']} skipped\n"
    )
    return col_map, counts


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mapping", required=True, help="Path to mapping JSON file")
    parser.add_argument("--excel", default=None, help="Override Excel filename")
    parser.add_argument("--db-name", default="macrostrat")
    parser.add_argument("--db-user", default="postgres")
    parser.add_argument("--db-password", default="")
    parser.add_argument("--db-host", default="localhost")
    parser.add_argument("--db-port", default="5432")
    parser.add_argument("--audit-dir", default="./audit", help="Directory for audit artifacts")
    parser.add_argument("--no-audit", action="store_true", help="Disable writing audit artifacts")
    args = parser.parse_args()

    mapping = load_mapping(args.mapping)
    sheets_map = mapping["sheets"]
    mapping_version = str(mapping["mapping_version"]).strip()

    excel_file = args.excel or mapping.get("excel_file_default") or "macrostrat_import.xlsx"
    excel_sha256 = sha256_file(excel_file)

    try:
        wb = load_workbook(excel_file, data_only=True)
    except FileNotFoundError:
        die(f"ERROR: File '{excel_file}' not found.")
    except Exception as e:
        die(f"ERROR: Could not read Excel file: {e}")

    meta_sheet = sheets_map["metadata"]
    excel_mapping_version = get_metadata_value(wb, meta_sheet, "mapping_version")
    if excel_mapping_version is None:
        die(f"ERROR: Excel metadata missing 'mapping_version'. Expected {mapping_version}")
    if str(excel_mapping_version).strip() != mapping_version:
        die(
            "ERROR: mapping_version mismatch.\n"
            f"  Excel metadata mapping_version: {str(excel_mapping_version).strip()}\n"
            f"  Mapping file mapping_version:   {mapping_version}"
        )

    project_id_val = get_metadata_value(wb, meta_sheet, "project_id")
    if project_id_val is None:
        die("ERROR: Excel metadata missing 'project_id'.")
    if isinstance(project_id_val, int):
        project_id = project_id_val
    elif isinstance(project_id_val, str) and project_id_val.isdigit():
        project_id = int(project_id_val)
    else:
        die("ERROR: 'project_id' must be an integer.")

    verify_project_id_via_api(project_id)

    conn = connect_db(
        dbname=args.db_name,
        user=args.db_user,
        password=args.db_password,
        host=args.db_host,
        port=args.db_port,
    )

    try:
        verify_project_id_in_db(conn, project_id)

        schema = SchemaInspector(conn)
        validate_mapping_against_schema(schema, mapping)

        ents = mapping["entities"]
        refs_ent = ents["refs"]
        col_groups_ent = ents["col_groups"]
        cols_ent = ents.get("cols")
        if cols_ent is None:
            die("ERROR: Mapping missing required entity 'cols' for Phase 2.5.")

        validate_cols_not_null_coverage(schema, cols_ent)

        # These will be populated inside the committed transaction and then used for audit writing.
        ref_map: Dict[str, int] = {}
        col_map: Dict[str, int] = {}
        refs_counts: Dict[str, int] = {}
        col_groups_counts: Dict[str, int] = {}
        cols_counts: Dict[str, Any] = {}

        with conn:
            ref_map, refs_counts = ingest_refs_mapping_driven(conn, wb, refs_ent, sheets_map)
            col_groups_map, col_groups_counts = ingest_col_groups_mapping_driven(conn, wb, col_groups_ent, sheets_map, project_id)
            cols_handlers = make_cols_handlers(col_groups_map)

            columns_sheet = sheets_map["columns"]
            col_map, cols_counts = ingest_columns_plugin(
                conn, wb, columns_sheet, project_id, ref_map, col_groups_map, cols_ent, cols_handlers
            )

        print("SUCCESS: Phase 2.5 ingestion completed (transaction committed).")

        if not args.no_audit:
            counts_manifest = {
                "refs": refs_counts,
                "col_groups": col_groups_counts,
                "cols": cols_counts,
                "units": {"processed": 0, "inserted": 0, "reused": 0, "skipped": True},
        }
        write_audit_artifacts(
            audit_dir=args.audit_dir,
            project_id=project_id,
            mapping_version=mapping_version,
            excel_file=os.path.basename(excel_file),
            excel_sha256=excel_sha256,
            ref_map=ref_map,
            col_map=col_map,
            counts=counts_manifest,
            unit_map=None,  # Phase 2.5: no units yet
        )


    except Exception as e:
        conn.rollback()
        die(f"ERROR: ingestion failed; rolled back transaction. Details: {e}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()